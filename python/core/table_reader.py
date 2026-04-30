import csv
import os
import sys
from datetime import datetime
from typing import Any, List, Optional, Tuple

try:
    from openpyxl import load_workbook as _openpyxl_load_workbook
except Exception:
    _openpyxl_load_workbook = None


def _sniff_delimiter(sample: str, preferred: Optional[str] = None) -> str:
    if preferred and len(preferred) == 1:
        return preferred

    sniff_candidates = [",", ";", "\t", "|"]
    if preferred and len(preferred) > 1:
        name = preferred.lower()
        mapping = {
            "comma": ",",
            ",": ",",
            "semicolon": ";",
            ";": ";",
            "tab": "\t",
            "\t": "\t",
            "pipe": "|",
            "|": "|",
            "auto": None,
        }
        mapped = mapping.get(name)
        if mapped is not None:
            return mapped

    try:
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample, delimiters="".join(sniff_candidates))
        return dialect.delimiter
    except Exception:
        counts = {d: sample.count(d) for d in sniff_candidates}
        best = max(counts, key=lambda k: counts[k])
        if counts[best] == 0:
            return ","
        return best


def _read_table(
    path: str,
    encoding: str = "utf-8-sig",
    delimiter: Optional[str] = None,
    xlsx_sheet: Optional[str] = None,
) -> Tuple[List[str], List[List[str]], str]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        if _openpyxl_load_workbook is None:
            raise RuntimeError(
                "XLSX input requires 'openpyxl'. Install it (e.g., pip install openpyxl) or provide CSV input."
            )

        wb = _openpyxl_load_workbook(path, data_only=True, read_only=True)
        try:
            if xlsx_sheet:
                if xlsx_sheet not in wb.sheetnames:
                    raise ValueError(
                        f"XLSX sheet '{xlsx_sheet}' not found. Available sheets: {wb.sheetnames}"
                    )
                ws = wb[xlsx_sheet]
            else:
                default_sheet_name = "data" if "data" in wb.sheetnames else wb.sheetnames[0]
                ws = wb[default_sheet_name]

            if delimiter and str(delimiter).lower() not in ("", "auto"):
                print(
                    "Warning: --delimiter is ignored for XLSX input.",
                    file=sys.stderr,
                )

            rows_iter = ws.iter_rows(values_only=True)
            first_row = next(rows_iter, None)
            if first_row is None:
                raise ValueError("XLSX appears to be empty.")

            def _to_text(cell: Any) -> str:
                if cell is None:
                    return ""
                if isinstance(cell, datetime):
                    return cell.isoformat()
                return str(cell)

            headers = [_to_text(c) for c in first_row]
            rows = [[_to_text(c) for c in row] for row in rows_iter]
            return headers, rows, f"xlsx:{ws.title}"
        finally:
            try:
                wb.close()
            except Exception:
                pass

    with open(path, "r", encoding=encoding, newline="") as f:
        sample = f.read(8192)
        f.seek(0)
        delim = _sniff_delimiter(sample, preferred=delimiter)
        reader = csv.reader(f, delimiter=delim)
        try:
            headers = next(reader)
        except StopIteration:
            raise ValueError("CSV appears to be empty.")
        rows = [list(r) for r in reader]
        return headers, rows, delim
