from __future__ import annotations

import os
import re
from typing import TYPE_CHECKING, Dict, Optional

if TYPE_CHECKING:
    from openpyxl import Workbook as WorkbookType
    from openpyxl.worksheet.worksheet import Worksheet as WorksheetType

try:
    from openpyxl.utils import get_column_letter as _get_column_letter
    from openpyxl.worksheet.table import Table as _Table
    from openpyxl.worksheet.table import TableStyleInfo as _TableStyleInfo
except Exception:
    _get_column_letter = None
    _Table = None
    _TableStyleInfo = None


def resolve_optional_file_path(raw_path: Optional[str]) -> Optional[str]:
    if raw_path is None:
        return None
    normalized = os.path.expanduser(raw_path.strip())
    return normalized or None


def read_theme_xml_bytes(theme_path_raw: str, source: str) -> bytes:
    theme_path = resolve_optional_file_path(theme_path_raw)
    if not theme_path:
        raise SystemExit(f"Empty XLSX theme file path from {source}")
    if not os.path.isfile(theme_path):
        raise SystemExit(f"XLSX theme file from {source} was not found: {theme_path}")
    try:
        with open(theme_path, "rb") as f:
            return f.read()
    except Exception as ex:
        raise SystemExit(
            f"Failed to read XLSX theme file from {source}: {theme_path} ({ex})"
        )


def apply_theme_to_workbook(
    workbook: "WorkbookType",
    *,
    theme_file: Optional[str] = None,
    default_theme_file: Optional[str] = None,
) -> None:
    """Apply workbook theme XML bytes from explicit or default file path.

    Priority:
    1) theme_file
    2) default_theme_file
    """
    if theme_file:
        workbook.loaded_theme = read_theme_xml_bytes(theme_file, "--xlsx-theme-file")
        return

    if default_theme_file and os.path.isfile(default_theme_file):
        workbook.loaded_theme = read_theme_xml_bytes(
            default_theme_file,
            "default theme",
        )


def _sanitize_table_name(name: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not safe:
        safe = "Sheet"
    if not re.match(r"^[A-Za-z_]", safe):
        safe = f"T_{safe}"
    return safe[:255]


def apply_table_style(
    worksheet: "WorksheetType",
    *,
    table_name: Optional[str] = None,
    style_name: str = "TableStyleMedium9",
) -> None:
    """Format worksheet used range as table with header row enabled."""
    if _Table is None or _TableStyleInfo is None or _get_column_letter is None:
        raise SystemExit(
            "XLSX table styling requires openpyxl. Install with: pip install openpyxl"
        )

    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return

    last_col = _get_column_letter(worksheet.max_column)
    table_ref = f"A1:{last_col}{worksheet.max_row}"
    style = _TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    existing_tables = list(worksheet.tables.values())
    if existing_tables:
        table = existing_tables[0]
        table.ref = table_ref
        table.tableStyleInfo = style
        return

    normalized_name = _sanitize_table_name(table_name or f"{worksheet.title}_Table")
    table = _Table(displayName=normalized_name, ref=table_ref)
    table.tableStyleInfo = style
    worksheet.add_table(table)


def autosize_columns(
    worksheet: "WorksheetType",
    *,
    min_width: float = 10.0,
    max_width: float = 60.0,
    padding: float = 2.0,
    manual_width_overrides: Optional[Dict[str, float]] = None,
) -> None:
    """Set column widths based on content length with optional per-column overrides."""
    if _get_column_letter is None:
        raise SystemExit(
            "XLSX column autosize requires openpyxl. Install with: pip install openpyxl"
        )

    lo = min(min_width, max_width)
    hi = max(min_width, max_width)
    overrides = {k.upper(): v for k, v in (manual_width_overrides or {}).items()}

    for col_idx in range(1, worksheet.max_column + 1):
        col_letter = _get_column_letter(col_idx)
        if col_letter in overrides:
            width = float(overrides[col_letter])
            worksheet.column_dimensions[col_letter].width = max(lo, min(hi, width))
            continue

        max_len = 0
        for row_idx in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            token = str(value)
            if len(token) > max_len:
                max_len = len(token)

        computed = float(max_len) + float(padding)
        worksheet.column_dimensions[col_letter].width = max(lo, min(hi, computed))
