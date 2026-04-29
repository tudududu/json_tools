#!/usr/bin/env python3
"""
SRT → CSV converter

Converts a SubRip (.srt) subtitle file into a simple CSV with columns:
"Start Time","End Time","Text"

Options:
- --fps <float>               Input frame rate for HH:MM:SS:FF output (default 25)
- --out-format <format>       Output format: 'frames' (HH:MM:SS:FF) or 'ms' (HH:MM:SS,SSS) [default: frames]
- --encoding <name>           Input file encoding (default: utf-8-sig)
- --quote-all                 Always quote all CSV fields (default OFF; minimal quoting otherwise)
- --delimiter <name>          CSV output delimiter: 'comma' or 'semicolon' (default 'comma')
- --output-type <csv|xlsx>    Output container override (otherwise inferred from output extension)
- --xlsx-theme-file <path>    Optional OOXML theme XML file to apply to generated XLSX workbooks
- --xlsx-template <path>      Optional XLSX template workbook to use as base for output

Examples:
    # Single file
    python python/tools/srt_to_csv.py input.srt output.csv --fps 25 --out-format frames
    python python/tools/srt_to_csv.py input.srt output.xlsx --fps 25 --out-format frames
    python python/tools/srt_to_csv.py input.srt output.csv --out-format ms

    # Batch mode: iterate all .srt in an input folder and write .csv to an output folder
    python python/tools/srt_to_csv.py --input-dir in_srt/ --output-dir out_csv/ --fps 25 --out-format frames
    python python/tools/srt_to_csv.py --input-dir in_srt/ --output-dir out_xlsx/ --output-type xlsx --fps 25 --out-format frames

    # Batch join: iterate all .srt in an input folder and write a single combined .csv/.xlsx
    python python/tools/srt_to_csv.py --input-dir in_srt/ joined.csv --join-output --fps 25 --out-format frames
    python python/tools/srt_to_csv.py --input-dir in_srt/ joined.xlsx --join-output --fps 25 --out-format frames
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from typing import TYPE_CHECKING, Callable, List, Optional, Tuple, Type

# openpyxl is an optional runtime dependency (only needed for XLSX output).
# The TYPE_CHECKING block gives Pylance accurate type information without
# requiring the package to be installed. The try/except block captures the
# actual runtime values; all three names are checked together before use.
if TYPE_CHECKING:
    from openpyxl import Workbook as WorkbookType
    from openpyxl.styles import Color as ColorType
    from openpyxl.styles import Font as FontType
    from openpyxl.styles import PatternFill as PatternFillType
    from openpyxl.worksheet.table import Table as TableType
    from openpyxl.worksheet.table import TableStyleInfo as TableStyleInfoType

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Color, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:  # pragma: no cover - optional dependency
    Workbook = None  # type: ignore[assignment,misc]
    load_workbook = None  # type: ignore[assignment,misc]
    Color = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]
    Table = None  # type: ignore[assignment,misc]
    TableStyleInfo = None  # type: ignore[assignment,misc]

_Workbook: Optional[Type["WorkbookType"]] = Workbook
_load_workbook: Optional[Callable[..., "WorkbookType"]] = load_workbook
_Color: Optional[Type["ColorType"]] = Color
_Font: Optional[Type["FontType"]] = Font
_PatternFill: Optional[Type["PatternFillType"]] = PatternFill
_Table: Optional[Type["TableType"]] = Table
_TableStyleInfo: Optional[Type["TableStyleInfoType"]] = TableStyleInfo

XLSX_TEMPLATE_ENV = "SRT_TO_CSV_XLSX_TEMPLATE"
XLSX_THEME_ENV = "SRT_TO_CSV_XLSX_THEME_FILE"
DEFAULT_XLSX_THEME_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "themes",
    "subtitles_theme.xml",
)


def _resolve_optional_file_path(raw_path: Optional[str]) -> Optional[str]:
    if raw_path is None:
        return None
    normalized = os.path.expanduser(raw_path.strip())
    return normalized or None


def _create_output_workbook(
    template_path_raw: Optional[str],
    template_source: str,
) -> "WorkbookType":
    """Create workbook, optionally from template, with explicit error handling.

    If a template is configured, loading must succeed; otherwise we
    abort with a clear message instead of silently falling back to default
    theme/workbook styling.
    """
    if _Workbook is None:
        raise SystemExit(
            "XLSX output requires openpyxl. Install with: pip install openpyxl"
        )

    template_path = _resolve_optional_file_path(template_path_raw)
    if not template_path:
        return _Workbook()

    if _load_workbook is None:
        raise SystemExit(
            "XLSX template loading requires openpyxl load_workbook support"
        )
    if not os.path.isfile(template_path):
        raise SystemExit(
            f"XLSX template path from {template_source} was not found: {template_path}"
        )

    try:
        return _load_workbook(template_path)
    except Exception as ex:
        raise SystemExit(
            f"Failed to load XLSX template from {template_source}: {template_path} ({ex})"
        )


def _read_theme_xml_bytes(theme_path_raw: str, source: str) -> bytes:
    theme_path = _resolve_optional_file_path(theme_path_raw)
    if not theme_path:
        raise SystemExit(f"Empty XLSX theme file path from {source}")
    if not os.path.isfile(theme_path):
        raise SystemExit(f"XLSX theme file from {source} was not found: {theme_path}")
    try:
        with open(theme_path, "rb") as f:
            return f.read()
    except Exception as ex:
        raise SystemExit(
            f"Failed to read XLSX theme file from {source}: {theme_path} ({ex})"
        )


def _resolve_theme_xml_bytes(
    template_path_raw: Optional[str],
    theme_path_raw: Optional[str],
) -> Optional[bytes]:
    """Resolve theme bytes for output workbook.

    Priority:
    1) --xlsx-theme-file
    2) SRT_TO_CSV_XLSX_THEME_FILE
    3) Built-in default theme file (only when no template workbook is used)
    """
    if theme_path_raw:
        return _read_theme_xml_bytes(theme_path_raw, "--xlsx-theme-file")

    env_theme_path = os.getenv(XLSX_THEME_ENV)
    if env_theme_path and env_theme_path.strip():
        return _read_theme_xml_bytes(env_theme_path, XLSX_THEME_ENV)

    # Preserve workbook theme from a provided template unless explicitly overridden.
    if _resolve_optional_file_path(template_path_raw):
        return None

    if os.path.isfile(DEFAULT_XLSX_THEME_FILE):
        return _read_theme_xml_bytes(DEFAULT_XLSX_THEME_FILE, "built-in theme")

    return None

TIME_RE = re.compile(
    r"^(?P<h1>\d{2}):(?P<m1>\d{2}):(?P<s1>\d{2})[,.](?P<ms1>\d{3})\s*-->\s*"
    r"(?P<h2>\d{2}):(?P<m2>\d{2}):(?P<s2>\d{2})[,.](?P<ms2>\d{3})\s*$"
)

HEADER = ["Start Time", "End Time", "Text"]
XLSX_HEADER = HEADER + [f"<ISO>{i}" for i in range(1, 11)]
FRAME_TC_RE = re.compile(r"^\d{2}:\d{2}:\d{2}:\d{2}$")
MS_TC_RE = re.compile(r"^\d{2}:\d{2}:\d{2}[,.]\d{3}$")


def parse_srt(lines: List[str]) -> List[Tuple[float, float, str]]:
    """Parse SRT lines into a list of (start_seconds, end_seconds, text)."""
    records: List[Tuple[float, float, str]] = []
    i = 0
    n = len(lines)
    while i < n:
        line = lines[i].strip("\r\n")
        m = TIME_RE.match(line)
        if not m:
            i += 1
            continue
        # Parse times
        h1 = int(m.group("h1"))
        m1 = int(m.group("m1"))
        s1 = int(m.group("s1"))
        ms1 = int(m.group("ms1"))
        h2 = int(m.group("h2"))
        m2 = int(m.group("m2"))
        s2 = int(m.group("s2"))
        ms2 = int(m.group("ms2"))
        start = h1 * 3600 + m1 * 60 + s1 + ms1 / 1000.0
        end = h2 * 3600 + m2 * 60 + s2 + ms2 / 1000.0
        # Collect following text lines until blank line
        i += 1
        text_lines: List[str] = []
        while i < n:
            lt = lines[i].rstrip("\r\n")
            if lt.strip() == "":
                break
            # Skip the optional numeric index line when encountered before time line; here we are past time line.
            text_lines.append(lt)
            i += 1
        # Move past the blank separator line (if present)
        while i < n and lines[i].strip() == "":
            i += 1
        text = "\n".join(text_lines)
        records.append((start, end, text))
    return records


def format_time_frames(seconds: float, fps: float) -> str:
    """Format time as HH:MM:SS:FF using nearest-frame rounding."""
    if fps <= 0:
        raise ValueError("fps must be > 0 for frames output")
    sec_int = int(seconds)
    frac = seconds - sec_int
    frames = int(round(frac * fps))
    # Carry over if frames equals fps due to rounding
    if frames >= int(fps):
        sec_int += 1
        frames -= int(fps)
    h = sec_int // 3600
    rem = sec_int % 3600
    m = rem // 60
    s = rem % 60
    return f"{h:02d}:{m:02d}:{s:02d}:{frames:02d}"


def format_time_ms(seconds: float) -> str:
    """Format time as HH:MM:SS,SSS with millisecond rounding."""
    sec_int = int(seconds)
    frac = seconds - sec_int
    ms = int(round(frac * 1000))
    if ms >= 1000:
        sec_int += 1
        ms -= 1000
    h = sec_int // 3600
    rem = sec_int % 3600
    m = rem // 60
    s = rem % 60
    return f"{h:02d}:{m:02d}:{s:02d},{ms:03d}"


def resolve_output_type(out_path: str, explicit_output_type: str | None) -> str:
    """Resolve output container from explicit flag or file extension."""
    if explicit_output_type:
        return explicit_output_type
    if out_path.lower().endswith(".xlsx"):
        return "xlsx"
    return "csv"


def records_to_rows(
    records: List[Tuple[float, float, str]], fps: float, out_format: str
) -> List[List[str]]:
    rows: List[List[str]] = []
    for start, end, text in records:
        if out_format == "frames":
            start_str = format_time_frames(start, fps)
            end_str = format_time_frames(end, fps)
        elif out_format == "ms":
            start_str = format_time_ms(start)
            end_str = format_time_ms(end)
        else:
            raise ValueError("out_format must be 'frames' or 'ms'")
        rows.append([start_str, end_str, text])
    return rows


def write_tabular_output(
    out_path: str,
    rows: List[List[str]],
    quote_all: bool,
    delimiter_name: str,
    output_type: str,
    xlsx_template: Optional[str] = None,
    xlsx_theme_file: Optional[str] = None,
) -> None:
    import os

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    if output_type == "csv":
        delimiter = "," if delimiter_name == "comma" else ";"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(
                f,
                delimiter=delimiter,
                quoting=(csv.QUOTE_ALL if quote_all else csv.QUOTE_MINIMAL),
            )
            writer.writerow(HEADER)
            writer.writerows(rows)
        return

    if output_type != "xlsx":
        raise ValueError("output_type must be 'csv' or 'xlsx'")
    # openpyxl symbols are checked together because XLSX formatting relies on
    # workbook, styles, and table objects.
    # Checking all needed symbols keeps the type narrowed to
    # non-None for the remainder of the XLSX path.
    if (
        _Workbook is None
        or _Color is None
        or _Font is None
        or _PatternFill is None
        or _Table is None
        or _TableStyleInfo is None
    ):
        raise SystemExit(
            "XLSX output requires openpyxl. Install with: pip install openpyxl"
        )

    resolved_template = xlsx_template
    template_source = "--xlsx-template"
    if not _resolve_optional_file_path(resolved_template):
        resolved_template = os.getenv(XLSX_TEMPLATE_ENV)
        template_source = XLSX_TEMPLATE_ENV
    wb = _create_output_workbook(resolved_template, template_source)
    theme_xml = _resolve_theme_xml_bytes(resolved_template, xlsx_theme_file)
    if theme_xml is not None:
        wb.loaded_theme = theme_xml

    ws = wb.active
    if ws is None:
        raise SystemExit("Failed to create XLSX worksheet")

    # Ensure the output worksheet starts from a clean slate while preserving
    # any workbook-level theme inherited from a template workbook.
    ws.delete_rows(1, ws.max_row)
    ws.tables.clear()
    ws.title = "subtitles"
    ws.append(XLSX_HEADER)

    # Requested widths are in pixels; map to nearest openpyxl column units.
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 41
    for col_letter in ("D", "E", "F", "G", "H", "I", "J", "K", "L", "M"):
        ws.column_dimensions[col_letter].width = 16

    # Excel theme color: Plum, Accent 5, Lighter 80%.
    title_fill = _PatternFill(fill_type="solid", fgColor=_Color(theme=8, tint=0.8))
    body_font = _Font(name="Aptos Narrow", size=12)

    for row in rows:
        normalized = list(row[: len(XLSX_HEADER)])
        if len(normalized) < len(XLSX_HEADER):
            normalized.extend([""] * (len(XLSX_HEADER) - len(normalized)))
        ws.append(normalized)

    # Apply body font to data rows only; keep header typography controlled by
    # the table style (for example white + bold in Medium styles).
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(XLSX_HEADER) + 1):
            ws.cell(row=row_idx, column=col_idx).font = body_font

    # Format the data range as a table with headers.
    if ws.max_row >= 2:
        table_ref = f"A1:M{ws.max_row}"
        existing_tables = list(ws.tables.values())
        if existing_tables:
            table = existing_tables[0]
            table.ref = table_ref
            table.tableStyleInfo = _TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
        else:
            table = _Table(displayName="SubtitlesTable", ref=table_ref)
            table.tableStyleInfo = _TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

    # Re-apply joined title row fill after table insertion so it remains visible.
    for row_idx in range(2, ws.max_row + 1):
        c1 = ws.cell(row=row_idx, column=1).value
        c2 = ws.cell(row=row_idx, column=2).value
        c3 = ws.cell(row=row_idx, column=3).value

        # Only style rows that contain at least one value in the data columns.
        in_use = any(value not in (None, "") for value in (c1, c2, c3))
        if not in_use:
            continue

        # Joined-output title marker rows are emitted as ["", "", <filename>].
        is_title_row = (
            c1 in (None, "") and c2 in (None, "") and str(c3 or "").strip() != ""
        )
        if is_title_row:
            for col_idx in range(1, len(XLSX_HEADER) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = title_fill

    wb.save(out_path)


def srt_to_csv(
    in_path: str,
    out_path: str,
    fps: float,
    out_format: str,
    encoding: str,
    quote_all: bool,
    delimiter_name: str,
    output_type: str | None = None,
    xlsx_template: str | None = None,
    xlsx_theme_file: str | None = None,
) -> None:
    with open(in_path, "r", encoding=encoding) as f:
        # Preserve lines; splitlines handles CRLF/CR/LF
        lines = f.read().splitlines()
    records = parse_srt(lines)
    rows = records_to_rows(records, fps=fps, out_format=out_format)
    resolved_output_type = resolve_output_type(out_path, output_type)
    write_tabular_output(
        out_path,
        rows,
        quote_all=quote_all,
        delimiter_name=delimiter_name,
        output_type=resolved_output_type,
        xlsx_template=xlsx_template,
        xlsx_theme_file=xlsx_theme_file,
    )


def _normalize_header_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (name or "").strip().lower())


def _resolve_column_index(
    headers: List[str],
    override: Optional[str],
    aliases: Tuple[str, ...],
) -> int:
    if override:
        if override.isdigit():
            idx = int(override) - 1
            if 0 <= idx < len(headers):
                return idx
            raise ValueError(f"Column index out of range: {override}")
        target = _normalize_header_name(override)
        for i, header in enumerate(headers):
            if _normalize_header_name(header) == target:
                return i
        raise ValueError(f"Column not found: {override}")

    normalized_headers = [_normalize_header_name(h) for h in headers]
    for alias in aliases:
        alias_norm = _normalize_header_name(alias)
        if alias_norm in normalized_headers:
            return normalized_headers.index(alias_norm)
    raise ValueError(f"Could not detect required column. Tried aliases: {aliases}")


def _read_reverse_table(
    in_path: str,
    encoding: str,
) -> Tuple[List[str], List[List[str]]]:
    if in_path.lower().endswith(".xlsx"):
        if _load_workbook is None:
            raise SystemExit(
                "XLSX input requires openpyxl. Install with: pip install openpyxl"
            )
        wb = _load_workbook(in_path, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("XLSX file has no active worksheet")
        all_rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(["" if v is None else str(v) for v in row])
        if not all_rows:
            return [], []
        headers = [c.strip() for c in all_rows[0]]
        return headers, all_rows[1:]

    with open(in_path, "r", newline="", encoding=encoding) as f:
        sample = f.read(8192)
        f.seek(0)
        delimiter = ","
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
            delimiter = dialect.delimiter
        except Exception:
            delimiter = ";" if ";" in sample and "," not in sample else ","
        reader = csv.reader(f, delimiter=delimiter)
        all_rows = [list(r) for r in reader]
    if not all_rows:
        return [], []
    headers = [(c or "").strip() for c in all_rows[0]]
    return headers, all_rows[1:]


def _detect_reverse_time_format(
    rows: List[List[str]],
    idx_start: int,
    idx_end: int,
) -> str:
    detected: Optional[str] = None
    for row in rows:
        start = (row[idx_start] if idx_start < len(row) else "").strip()
        end = (row[idx_end] if idx_end < len(row) else "").strip()
        if not start and not end:
            continue
        if not start or not end:
            raise ValueError(
                "Row has only one time value; both Start Time and End Time are required"
            )
        for token in (start, end):
            current: Optional[str] = None
            if FRAME_TC_RE.fullmatch(token):
                current = "frames"
            elif MS_TC_RE.fullmatch(token):
                current = "ms"
            else:
                raise ValueError(f"Unsupported timecode format: {token}")
            if detected is None:
                detected = current
            elif detected != current:
                raise ValueError("Mixed timecode formats detected in a single file")
    if detected is None:
        raise ValueError("No valid timed subtitle rows found in input")
    return detected


def _parse_reverse_timecode(value: str, fmt: str, fps: float) -> float:
    token = value.strip()
    if fmt == "frames":
        h, m, s, ff = [int(part) for part in token.split(":")]
        if fps <= 0:
            raise ValueError("fps must be > 0 for frames input")
        return h * 3600 + m * 60 + s + ff / fps
    if fmt == "ms":
        base, ms = re.split(r"[,.]", token, maxsplit=1)
        h, m, s = [int(part) for part in base.split(":")]
        return h * 3600 + m * 60 + s + int(ms) / 1000.0
    raise ValueError(f"Unsupported timecode format: {fmt}")


def _rows_to_reverse_records(
    rows: List[List[str]],
    idx_start: int,
    idx_end: int,
    idx_text: int,
    time_format: str,
    fps: float,
) -> List[Tuple[float, float, str]]:
    out: List[Tuple[float, float, str]] = []
    for row in rows:
        start = (row[idx_start] if idx_start < len(row) else "").strip()
        end = (row[idx_end] if idx_end < len(row) else "").strip()
        text = row[idx_text] if idx_text < len(row) else ""

        # Skip joined marker rows emitted by --join-output in forward mode.
        if not start and not end:
            continue
        if not start or not end:
            raise ValueError(
                "Row has only one time value; both Start Time and End Time are required"
            )

        tin = _parse_reverse_timecode(start, time_format, fps)
        tout = _parse_reverse_timecode(end, time_format, fps)
        out.append((tin, tout, text))
    return out


def _records_to_srt_text(records: List[Tuple[float, float, str]]) -> str:
    lines: List[str] = []
    for idx, (start, end, text) in enumerate(records, start=1):
        lines.append(str(idx))
        lines.append(f"{format_time_ms(start)} --> {format_time_ms(end)}")
        text_lines = text.splitlines() if text else [""]
        lines.extend(text_lines)
        lines.append("")
    return "\n".join(lines)


def _extract_joined_reverse_blocks(
    rows: List[List[str]],
    idx_start: int,
    idx_end: int,
    idx_text: int,
) -> List[Tuple[str, List[List[str]]]]:
    blocks: List[Tuple[str, List[List[str]]]] = []
    current_name: Optional[str] = None
    current_rows: List[List[str]] = []
    marker_count = 0

    for row in rows:
        start = (row[idx_start] if idx_start < len(row) else "").strip()
        end = (row[idx_end] if idx_end < len(row) else "").strip()
        text = (row[idx_text] if idx_text < len(row) else "").strip()

        if not start and not end:
            if not text:
                continue
            marker_count += 1
            if current_name is not None:
                blocks.append((current_name, current_rows))
            current_name = text
            current_rows = []
            continue

        if not start or not end:
            raise ValueError(
                "Row has only one time value; both Start Time and End Time are required"
            )
        if current_name is None:
            raise ValueError(
                "Joined reverse input requires marker rows before timed rows"
            )
        current_rows.append(row)

    if current_name is not None:
        blocks.append((current_name, current_rows))

    if marker_count == 0:
        raise ValueError(
            "Joined reverse input requires marker rows (empty Start/End with filename in Text)"
        )

    return blocks


def _sanitize_joined_marker_filename(marker: str) -> str:
    raw = (marker or "").strip()
    raw = re.sub(r"\.srt$", "", raw, flags=re.I)
    raw = re.sub(r"[\\/]+", "_", raw)
    raw = re.sub(r"[^A-Za-z0-9._ -]", "_", raw)
    raw = re.sub(r"\s+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("._ ")
    if not raw:
        raw = "subtitle"
    return f"{raw}.srt"


def _dedupe_output_filename(name: str, used: set[str]) -> str:
    base, ext = os.path.splitext(name)
    candidate = name
    i = 2
    while candidate.lower() in used:
        candidate = f"{base}_{i}{ext}"
        i += 1
    used.add(candidate.lower())
    return candidate


def csv_to_srt(
    in_path: str,
    out_path: str,
    fps: float,
    encoding: str,
    start_col: Optional[str] = None,
    end_col: Optional[str] = None,
    text_col: Optional[str] = None,
) -> None:
    import os

    headers, rows = _read_reverse_table(in_path, encoding=encoding)
    if not headers:
        raise ValueError(f"Input table is empty: {in_path}")

    idx_start = _resolve_column_index(
        headers,
        start_col,
        aliases=("Start Time", "start", "in", "inpoint"),
    )
    idx_end = _resolve_column_index(
        headers,
        end_col,
        aliases=("End Time", "end", "out", "outpoint"),
    )
    idx_text = _resolve_column_index(
        headers,
        text_col,
        aliases=("Text", "subtitle", "caption"),
    )

    time_format = _detect_reverse_time_format(rows, idx_start, idx_end)
    records = _rows_to_reverse_records(
        rows,
        idx_start=idx_start,
        idx_end=idx_end,
        idx_text=idx_text,
        time_format=time_format,
        fps=fps,
    )

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with open(out_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(_records_to_srt_text(records))


def csv_to_srt_joined(
    in_path: str,
    out_dir: str,
    fps: float,
    encoding: str,
    start_col: Optional[str] = None,
    end_col: Optional[str] = None,
    text_col: Optional[str] = None,
) -> List[str]:
    headers, rows = _read_reverse_table(in_path, encoding=encoding)
    if not headers:
        raise ValueError(f"Input table is empty: {in_path}")

    idx_start = _resolve_column_index(
        headers,
        start_col,
        aliases=("Start Time", "start", "in", "inpoint"),
    )
    idx_end = _resolve_column_index(
        headers,
        end_col,
        aliases=("End Time", "end", "out", "outpoint"),
    )
    idx_text = _resolve_column_index(
        headers,
        text_col,
        aliases=("Text", "subtitle", "caption"),
    )

    blocks = _extract_joined_reverse_blocks(rows, idx_start, idx_end, idx_text)
    timed_rows = [r for _, block_rows in blocks for r in block_rows]
    time_format = _detect_reverse_time_format(timed_rows, idx_start, idx_end)

    os.makedirs(out_dir, exist_ok=True)
    used_names: set[str] = set()
    written: List[str] = []

    for marker_name, block_rows in blocks:
        if not block_rows:
            continue
        records = _rows_to_reverse_records(
            block_rows,
            idx_start=idx_start,
            idx_end=idx_end,
            idx_text=idx_text,
            time_format=time_format,
            fps=fps,
        )
        if not records:
            continue
        fname = _sanitize_joined_marker_filename(marker_name)
        fname = _dedupe_output_filename(fname, used_names)
        out_path = os.path.join(out_dir, fname)
        with open(out_path, "w", encoding="utf-8", newline="\n") as f:
            f.write(_records_to_srt_text(records))
        written.append(out_path)

    if not written:
        raise ValueError("No timed subtitle blocks found under joined markers")
    return written


def main() -> None:
    p = argparse.ArgumentParser(description="Convert SRT<->CSV/XLSX subtitles")
    p.add_argument(
        "input", nargs="?", help="Path to input .srt file (omit when using --input-dir)"
    )
    p.add_argument(
        "output",
        nargs="?",
        help="Path to output .csv/.xlsx file (omit when using --output-dir; in --join-output mode this can be the joined output file path)",
    )
    p.add_argument(
        "--input-dir", help="Directory containing .srt files to convert (batch mode)"
    )
    p.add_argument("--output-dir", help="Directory to write .csv files in batch mode")
    p.add_argument(
        "--join-output",
        action="store_true",
        help="In batch mode, join all inputs into a single output CSV (provide output file path)",
    )
    p.add_argument(
        "--reverse",
        action="store_true",
        help="Reverse mode: convert CSV/XLSX back to SRT",
    )
    p.add_argument(
        "--reverse-joined",
        action="store_true",
        help="Reverse mode: parse joined CSV/XLSX input (marker rows) and split to multiple SRT files",
    )
    p.add_argument(
        "--fps",
        type=float,
        default=25.0,
        help="Input frame rate for frames output (default 25)",
    )
    p.add_argument(
        "--out-format",
        choices=["frames", "ms"],
        default="frames",
        help="Output format: frames (HH:MM:SS:FF) or ms (HH:MM:SS,SSS)",
    )
    p.add_argument(
        "--encoding",
        default="utf-8-sig",
        help="Input file encoding (default utf-8-sig)",
    )
    p.add_argument(
        "--quote-all",
        action="store_true",
        help="Always quote all CSV fields (default OFF)",
    )
    p.add_argument(
        "--delimiter",
        choices=["comma", "semicolon"],
        default="comma",
        help="Output delimiter (default comma)",
    )
    p.add_argument(
        "--output-type",
        choices=["csv", "xlsx"],
        help="Output container override. When omitted, inferred from output extension (.xlsx => xlsx, otherwise csv)",
    )
    p.add_argument(
        "--xlsx-template",
        default=None,
        help=(
            "Optional path to an XLSX template workbook used as output base. "
            f"If omitted, falls back to {XLSX_TEMPLATE_ENV} env var when set."
        ),
    )
    p.add_argument(
        "--xlsx-theme-file",
        default=None,
        help=(
            "Optional OOXML theme XML file applied to generated XLSX. "
            f"If omitted, falls back to {XLSX_THEME_ENV} env var; otherwise a built-in theme is used when available."
        ),
    )
    p.add_argument(
        "--start-col",
        help="Reverse mode: start time column name or 1-based index",
    )
    p.add_argument(
        "--end-col",
        help="Reverse mode: end time column name or 1-based index",
    )
    p.add_argument(
        "--text-col",
        help="Reverse mode: text column name or 1-based index",
    )
    args = p.parse_args()

    # Determine mode: single file or batch directory
    import os

    if args.reverse or args.reverse_joined:
        if args.join_output:
            raise SystemExit("--join-output is not supported in reverse mode yet")

        if args.reverse_joined:
            if args.output:
                raise SystemExit(
                    "--reverse-joined writes multiple files; use --output-dir (or omit to use input directory)"
                )
            if args.input_dir:
                in_dir = args.input_dir
                if not os.path.isdir(in_dir):
                    raise SystemExit(f"Input directory not found: {in_dir}")
                names = [
                    n
                    for n in sorted(os.listdir(in_dir))
                    if n.lower().endswith(".csv") or n.lower().endswith(".xlsx")
                ]
                out_dir = args.output_dir or args.input_dir
                os.makedirs(out_dir, exist_ok=True)
                for name in names:
                    in_path = os.path.join(in_dir, name)
                    csv_to_srt_joined(
                        in_path,
                        out_dir,
                        fps=args.fps,
                        encoding=args.encoding,
                        start_col=args.start_col,
                        end_col=args.end_col,
                        text_col=args.text_col,
                    )
            else:
                if not args.input:
                    raise SystemExit(
                        "Provide input path, or use --input-dir/--output-dir for batch mode"
                    )
                if not os.path.isfile(args.input):
                    raise SystemExit(f"No such file or directory: '{args.input}'")
                out_dir = args.output_dir or os.path.dirname(args.input) or "."
                csv_to_srt_joined(
                    args.input,
                    out_dir,
                    fps=args.fps,
                    encoding=args.encoding,
                    start_col=args.start_col,
                    end_col=args.end_col,
                    text_col=args.text_col,
                )
            return

        if args.input_dir:
            in_dir = args.input_dir
            if not os.path.isdir(in_dir):
                raise SystemExit(f"Input directory not found: {in_dir}")
            names = [
                n
                for n in sorted(os.listdir(in_dir))
                if n.lower().endswith(".csv") or n.lower().endswith(".xlsx")
            ]
            out_dir = args.output_dir or args.input_dir
            os.makedirs(out_dir, exist_ok=True)
            for name in names:
                in_path = os.path.join(in_dir, name)
                base = os.path.splitext(name)[0]
                out_path = os.path.join(out_dir, base + ".srt")
                csv_to_srt(
                    in_path,
                    out_path,
                    fps=args.fps,
                    encoding=args.encoding,
                    start_col=args.start_col,
                    end_col=args.end_col,
                    text_col=args.text_col,
                )
        else:
            if not args.input or not args.output:
                raise SystemExit(
                    "Provide input and output paths, or use --input-dir/--output-dir for batch mode"
                )
            if not os.path.isfile(args.input):
                raise SystemExit(f"No such file or directory: '{args.input}'")
            csv_to_srt(
                args.input,
                args.output,
                fps=args.fps,
                encoding=args.encoding,
                start_col=args.start_col,
                end_col=args.end_col,
                text_col=args.text_col,
            )
        return

    if args.input_dir:
        in_dir = args.input_dir
        if not os.path.isdir(in_dir):
            raise SystemExit(f"Input directory not found: {in_dir}")
        names = [n for n in sorted(os.listdir(in_dir)) if n.lower().endswith(".srt")]
        if args.join_output:
            # Resolve output file path: allow positional output OR positional input token used as output OR --output-dir as file path
            out_path = None
            if args.output:
                out_path = args.output
            elif args.input:
                # Users might place the output file path in the positional 'input' when using --input-dir
                out_path = args.input
            elif args.output_dir:
                out_path = args.output_dir
            if not out_path:
                raise SystemExit(
                    "Provide an output file path for --join-output mode (positional after --input-dir or via --output-dir)"
                )
            rows: List[List[str]] = []
            for name in names:
                in_path = os.path.join(in_dir, name)
                rows.append(["", "", name])
                with open(in_path, "r", encoding=args.encoding) as sf:
                    lines = sf.read().splitlines()
                rows.extend(
                    records_to_rows(
                        parse_srt(lines), fps=args.fps, out_format=args.out_format
                    )
                )
            resolved_output_type = resolve_output_type(out_path, args.output_type)
            write_tabular_output(
                out_path,
                rows,
                quote_all=args.quote_all,
                delimiter_name=args.delimiter,
                output_type=resolved_output_type,
                xlsx_template=args.xlsx_template,
                xlsx_theme_file=args.xlsx_theme_file,
            )
        else:
            out_dir = args.output_dir or args.input_dir
            os.makedirs(out_dir, exist_ok=True)
            for name in names:
                in_path = os.path.join(in_dir, name)
                base = os.path.splitext(name)[0]
                ext = ".xlsx" if args.output_type == "xlsx" else ".csv"
                out_name = base + ext
                out_path = os.path.join(out_dir, out_name)
                srt_to_csv(
                    in_path,
                    out_path,
                    fps=args.fps,
                    out_format=args.out_format,
                    encoding=args.encoding,
                    quote_all=args.quote_all,
                    delimiter_name=args.delimiter,
                    output_type=args.output_type,
                    xlsx_template=args.xlsx_template,
                    xlsx_theme_file=args.xlsx_theme_file,
                )
    else:
        if not args.input or not args.output:
            raise SystemExit(
                "Provide input and output paths, or use --input-dir/--output-dir for batch mode"
            )
        if not os.path.isfile(args.input):
            raise SystemExit(f"No such file or directory: '{args.input}'")
        srt_to_csv(
            args.input,
            args.output,
            fps=args.fps,
            out_format=args.out_format,
            encoding=args.encoding,
            quote_all=args.quote_all,
            delimiter_name=args.delimiter,
            output_type=args.output_type,
            xlsx_template=args.xlsx_template,
            xlsx_theme_file=args.xlsx_theme_file,
        )


if __name__ == "__main__":
    main()
