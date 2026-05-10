#!/usr/bin/env python3
"""
Generate an Excel template for LAYER_NAME_CONFIG from a sample JSON file.

Output workbook shape:
- Sheet "LAYER_NAME_CONFIG_items" with columns: key, exact, contains
- Sheet "LAYER_NAME_CONFIG_recenterRules" with columns: force, noRecenter, alignH, alignV
- Sheet "TIMING_BEHAVIOR" (optional) with columns: layerName, behavior
- Sheet "TIMING_ITEM_SELECTOR" (optional) with columns: itemName, mode, value
- Sheet "SKIP_COPY_CONFIG" (optional) with columns: key, value, names
- Sheet "MODULE_MAP" (optional) with columns: module, ENABLED, SOURCE_KEY
- Sheet "EXPLICIT_VARIANTS_BY_VIDEOID" (optional) with columns: video_id, variants
"""

from __future__ import annotations

import argparse
import json
import os
from typing import TYPE_CHECKING, Dict, List, Optional, Sequence, cast

from .sheet_names_config import SHEETS_BY_KEY

try:
    from .xlsx_styling import (
        apply_table_style,
        apply_theme_to_workbook,
        autosize_columns,
    )
except ImportError:
    from python.tools.xlsx_styling import (
        apply_table_style,
        apply_theme_to_workbook,
        autosize_columns,
    )

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet as WorksheetType

try:
    from openpyxl import Workbook as _Workbook
    from openpyxl.worksheet.datavalidation import DataValidation as _DataValidation
except Exception:
    _Workbook = None
    _DataValidation = None

RE_CENTER_KEYS: Sequence[str] = ("force", "noRecenter", "alignH", "alignV")
VALID_SELECTOR_MODES: Sequence[str] = ("line", "index", "minMax")
DEFAULT_XLSX_THEME_FILE = os.path.normpath(
    os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "themes",
        "subtitles_theme.xml",
    )
)


def _to_list(value: object) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    token = str(value).strip()
    return [token] if token else []


def generate_template(
    input_json: str,
    output_xlsx: str,
    separator: str,
    root_key: str,
    xlsx_theme_file: Optional[str] = None,
    min_column_width: float = 10.0,
    max_column_width: float = 60.0,
    column_width_overrides_by_sheet: Optional[Dict[str, Dict[str, float]]] = None,
) -> None:
    if _Workbook is None:
        raise RuntimeError(
            "XLSX support requires openpyxl. Install with: pip install openpyxl"
        )

    with open(input_json, "r", encoding="utf-8") as f:
        raw = json.load(f)

    body: Optional[Dict[str, object]] = None
    timing_behavior_map: Optional[Dict[str, object]] = None
    timing_item_selector_map: Optional[Dict[str, object]] = None
    skip_copy_config_map: Optional[Dict[str, object]] = None
    explicit_variants_by_videoid_map: Optional[Dict[str, object]] = None
    module_map_map: Optional[Dict[str, object]] = None

    # Parse config structure: config.addLayers and config.modular
    config = raw.get("config")
    if isinstance(config, dict):
        add_layers = config.get("addLayers")
        if isinstance(add_layers, dict):
            maybe_body = add_layers.get(root_key)
            if isinstance(maybe_body, dict):
                body = maybe_body
            tb_raw = add_layers.get("TIMING_BEHAVIOR")
            if isinstance(tb_raw, dict):
                timing_behavior_map = tb_raw
            tis_raw = add_layers.get("TIMING_ITEM_SELECTOR")
            if isinstance(tis_raw, dict):
                timing_item_selector_map = tis_raw
            scc_raw = add_layers.get("SKIP_COPY_CONFIG")
            if isinstance(scc_raw, dict):
                skip_copy_config_map = scc_raw
        modular = config.get("modular")
        if isinstance(modular, dict):
            maybe_body = modular.get(root_key)
            if isinstance(maybe_body, dict):
                body = maybe_body
            evbv_raw = modular.get("EXPLICIT_VARIANTS_BY_VIDEOID")
            if isinstance(evbv_raw, dict):
                explicit_variants_by_videoid_map = evbv_raw
            mm_raw = modular.get("MODULE_MAP")
            if isinstance(mm_raw, dict):
                module_map_map = mm_raw

    if body is None:
        raise ValueError(
            f"Root key '{root_key}' not found in config.addLayers or config.modular"
        )

    layer_names_sheet = SHEETS_BY_KEY["LAYER_NAME_CONFIG_items"].default_sheet_name
    recenter_rules_sheet = SHEETS_BY_KEY[
        "LAYER_NAME_CONFIG_recenterRules"
    ].default_sheet_name

    wb = _Workbook()
    ws_layers = cast("WorksheetType", wb.active)
    created_sheets: List["WorksheetType"] = [ws_layers]
    ws_layers.title = layer_names_sheet
    ws_layers.append(["key", "exact", "contains"])

    for key, value in body.items():
        if key == "recenterRules":
            continue
        if not isinstance(value, dict):
            continue

        exact = _to_list(value.get("exact"))
        contains = _to_list(value.get("contains"))
        ws_layers.append([key, separator.join(exact), separator.join(contains)])

    ws_rules = wb.create_sheet(title=recenter_rules_sheet)
    created_sheets.append(ws_rules)
    ws_rules.append(list(RE_CENTER_KEYS))

    recenter = body.get("recenterRules")
    if isinstance(recenter, dict):
        columns = {k: _to_list(recenter.get(k)) for k in RE_CENTER_KEYS}
    else:
        columns = {k: [] for k in RE_CENTER_KEYS}

    max_rows = max((len(columns[k]) for k in RE_CENTER_KEYS), default=0)
    for i in range(max_rows):
        ws_rules.append(
            [
                columns["force"][i] if i < len(columns["force"]) else "",
                columns["noRecenter"][i] if i < len(columns["noRecenter"]) else "",
                columns["alignH"][i] if i < len(columns["alignH"]) else "",
                columns["alignV"][i] if i < len(columns["alignV"]) else "",
            ]
        )

    if timing_behavior_map is not None:
        ws_tb = wb.create_sheet(
            title=SHEETS_BY_KEY["TIMING_BEHAVIOR"].default_sheet_name
        )
        created_sheets.append(ws_tb)
        ws_tb.append(["layerName", "behavior"])
        for layer_name, behavior in timing_behavior_map.items():
            ws_tb.append([str(layer_name), str(behavior)])

        if _DataValidation is not None:
            dv = _DataValidation(
                type="list", formula1='"timed,span,asIs"', allow_blank=True
            )
            ws_tb.add_data_validation(dv)
            # Lock behavior values in column B for template editing.
            dv.add(f"B2:B{max(ws_tb.max_row, 2)}")

    if timing_item_selector_map is not None:
        ws_tis = wb.create_sheet(
            title=SHEETS_BY_KEY["TIMING_ITEM_SELECTOR"].default_sheet_name
        )
        created_sheets.append(ws_tis)
        ws_tis.append(["itemName", "mode", "value"])
        for item_name, config_value in timing_item_selector_map.items():
            if not isinstance(config_value, dict):
                continue
            mode = config_value.get("mode", "")
            value = config_value.get("value", "")
            ws_tis.append([str(item_name), str(mode), value])

        if _DataValidation is not None:
            dv = _DataValidation(
                type="list", formula1='"line,index,minMax"', allow_blank=True
            )
            ws_tis.add_data_validation(dv)
            dv.add(f"B2:B{max(ws_tis.max_row, 2)}")

    if skip_copy_config_map is not None:
        ws_scc = wb.create_sheet(
            title=SHEETS_BY_KEY["SKIP_COPY_CONFIG"].default_sheet_name
        )
        created_sheets.append(ws_scc)
        ws_scc.append(["key", "value", "names"])
        for key, config_value in skip_copy_config_map.items():
            if isinstance(config_value, dict):
                names_raw = config_value.get("names")
                if names_raw is None:
                    names_raw = config_value.get("keys")
                if names_raw is None:
                    names_raw = config_value.get("tokens")
                names = _to_list(names_raw)
                enabled_raw = config_value.get("enabled")
                enabled = bool(enabled_raw) if enabled_raw is not None else bool(names)
                ws_scc.append([str(key), enabled, separator.join(names)])
            elif isinstance(config_value, list):
                names = _to_list(config_value)
                ws_scc.append([str(key), bool(names), separator.join(names)])
            else:
                ws_scc.append([str(key), bool(config_value), ""])

        if _DataValidation is not None:
            dv = _DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
            ws_scc.add_data_validation(dv)
            dv.add(f"B2:B{max(ws_scc.max_row, 2)}")

    if explicit_variants_by_videoid_map is not None:
        ws_evbv = wb.create_sheet(
            title=SHEETS_BY_KEY["EXPLICIT_VARIANTS_BY_VIDEOID"].default_sheet_name
        )
        created_sheets.append(ws_evbv)
        ws_evbv.append(["video_id", "variants"])
        for video_id, variants in explicit_variants_by_videoid_map.items():
            variants_str = separator.join(_to_list(variants))
            ws_evbv.append([str(video_id), variants_str])

    if module_map_map is not None:
        ws_mm = wb.create_sheet(title=SHEETS_BY_KEY["MODULE_MAP"].default_sheet_name)
        created_sheets.append(ws_mm)
        ws_mm.append(["module", "ENABLED", "SOURCE_KEY"])
        for module_name, config_keys in module_map_map.items():
            if not isinstance(config_keys, dict):
                continue
            mode = config_keys.get("ENABLED", "")
            value = config_keys.get("SOURCE_KEY", "")
            ws_mm.append([str(module_name), bool(mode), value])

        if _DataValidation is not None:
            dv = _DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
            ws_mm.add_data_validation(dv)
            dv.add(f"B2:B{max(ws_mm.max_row, 2)}")

    apply_theme_to_workbook(
        wb,
        theme_file=xlsx_theme_file,
        default_theme_file=DEFAULT_XLSX_THEME_FILE,
    )

    overrides_by_sheet = column_width_overrides_by_sheet or {}
    for idx, ws in enumerate(created_sheets, start=1):
        apply_table_style(
            ws,
            table_name=f"TemplateTable_{idx}",
            style_name="TableStyleMedium9",
        )
        autosize_columns(
            ws,
            min_width=min_column_width,
            max_width=max_column_width,
            manual_width_overrides=overrides_by_sheet.get(ws.title),
        )

    os.makedirs(os.path.dirname(output_xlsx) or ".", exist_ok=True)
    wb.save(output_xlsx)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate LAYER_NAME_CONFIG Excel template from JSON sample"
    )
    parser.add_argument("input", help="Path to input JSON sample")
    parser.add_argument("output", help="Path to output XLSX template")
    parser.add_argument(
        "--separator",
        default="; ",
        help="Separator used when writing list cells (default '; ')",
    )
    parser.add_argument(
        "--root-key",
        default="LAYER_NAME_CONFIG",
        help="Root key in input JSON (default LAYER_NAME_CONFIG)",
    )
    parser.add_argument(
        "--xlsx-theme-file",
        default=None,
        help=(
            "Optional path to workbook theme XML file. "
            "If omitted, built-in python/tools/themes/subtitles_theme.xml is used when present."
        ),
    )
    parser.add_argument(
        "--min-column-width",
        type=float,
        default=10.0,
        help="Minimum dynamic column width (default 10.0)",
    )
    parser.add_argument(
        "--max-column-width",
        type=float,
        default=60.0,
        help="Maximum dynamic column width (default 60.0)",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        raise SystemExit(f"No such file or directory: '{args.input}'")

    generate_template(
        input_json=args.input,
        output_xlsx=args.output,
        separator=args.separator,
        root_key=args.root_key,
        xlsx_theme_file=args.xlsx_theme_file,
        min_column_width=args.min_column_width,
        max_column_width=args.max_column_width,
    )


if __name__ == "__main__":
    main()
