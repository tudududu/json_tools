#!/usr/bin/env python3
"""
XLSX -> LAYER_NAME_CONFIG JSON converter.

Expected workbook shape:
- Sheet "LayerNames" with columns: key, exact, contains
- Sheet "RecenterRules" with columns: force, noRecenter, alignH, alignV

Rules:
- Sheet names are matched case-insensitively.
- exact/contains are split only by the explicit --separator argument.
- exact/contains are always emitted as arrays, even when empty.
"""

from __future__ import annotations

import argparse
import json
import os
from typing import TYPE_CHECKING, Dict, Iterable, List, Sequence, cast

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet as WorksheetType

try:
    from openpyxl import load_workbook as _openpyxl_load_workbook
except Exception:
    _openpyxl_load_workbook = None

RE_CENTER_KEYS: Sequence[str] = ("force", "noRecenter", "alignH", "alignV")


def _norm_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def _sheet_by_name_ci(workbook: object, configured_name: str) -> "WorksheetType":
    target = configured_name.strip().lower()
    sheets = getattr(workbook, "worksheets", [])
    for ws in sheets:
        if str(getattr(ws, "title", "")).strip().lower() == target:
            return ws
    raise ValueError(f"Sheet not found (case-insensitive): {configured_name}")


def _split_list_cell(value: object, separator: str) -> List[str]:
    token = str(value or "").strip()
    if not token:
        return []
    return [part.strip() for part in token.split(separator) if part.strip()]


def _read_headers(worksheet: "WorksheetType") -> List[str]:
    row_iter = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
    first = next(row_iter, None)
    if first is None:
        raise ValueError("Worksheet is empty (missing header row)")
    return [str(c or "").strip() for c in first]


def _index_map(headers: Iterable[str]) -> Dict[str, int]:
    return {_norm_header(name): idx for idx, name in enumerate(headers)}


def _cell(row: Sequence[object], idx: int) -> object:
    return row[idx] if idx < len(row) else ""


def _parse_layer_names(
    worksheet: "WorksheetType", separator: str
) -> Dict[str, Dict[str, List[str]]]:
    headers = _read_headers(worksheet)
    idx = _index_map(headers)

    for required in ("key", "exact", "contains"):
        if required not in idx:
            raise ValueError(f"LayerNames sheet is missing required column: {required}")

    out: Dict[str, Dict[str, List[str]]] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        key = str(_cell(row, idx["key"]) or "").strip()
        if not key:
            continue

        exact = _split_list_cell(_cell(row, idx["exact"]), separator)
        contains = _split_list_cell(_cell(row, idx["contains"]), separator)
        out[key] = {"exact": exact, "contains": contains}

    return out


def _parse_recenter_rules(worksheet: "WorksheetType") -> Dict[str, List[str]]:
    headers = _read_headers(worksheet)
    idx = _index_map(headers)

    for required in RE_CENTER_KEYS:
        if _norm_header(required) not in idx:
            raise ValueError(
                f"RecenterRules sheet is missing required column: {required}"
            )

    out: Dict[str, List[str]] = {k: [] for k in RE_CENTER_KEYS}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for key in RE_CENTER_KEYS:
            col_idx = idx[_norm_header(key)]
            value = str(_cell(row, col_idx) or "").strip()
            if value:
                out[key].append(value)

    return out


def convert_workbook(
    in_path: str,
    separator: str,
    layer_names_sheet: str,
    recenter_rules_sheet: str,
    root_key: str,
) -> Dict[str, Dict[str, object]]:
    if _openpyxl_load_workbook is None:
        raise RuntimeError(
            "XLSX support requires openpyxl. Install with: pip install openpyxl"
        )

    wb = _openpyxl_load_workbook(in_path, read_only=True, data_only=True)
    try:
        ws_layers = _sheet_by_name_ci(wb, layer_names_sheet)
        ws_rules = _sheet_by_name_ci(wb, recenter_rules_sheet)

        layer_map = _parse_layer_names(ws_layers, separator)
        rules_map = _parse_recenter_rules(ws_rules)

        body: Dict[str, object] = dict(layer_map)
        body["recenterRules"] = rules_map
        return {root_key: body}
    finally:
        wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert LAYER_NAME_CONFIG Excel workbook to JSON"
    )
    parser.add_argument("input", help="Path to input XLSX")
    parser.add_argument("output", help="Path to output JSON")
    parser.add_argument(
        "--separator",
        default=";",
        help="Explicit separator for exact/contains cells (default ';')",
    )
    parser.add_argument(
        "--layer-names-sheet",
        default="LayerNames",
        help="Layer names sheet name (case-insensitive match, default LayerNames)",
    )
    parser.add_argument(
        "--recenter-rules-sheet",
        default="RecenterRules",
        help="Recenter rules sheet name (case-insensitive match, default RecenterRules)",
    )
    parser.add_argument(
        "--root-key",
        default="LAYER_NAME_CONFIG",
        help="Root key in output JSON (default LAYER_NAME_CONFIG)",
    )
    parser.add_argument(
        "--indent",
        type=int,
        default=4,
        help="JSON indentation (default 4; set 0 for compact)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and print summary only (no output file)",
    )
    args = parser.parse_args()

    if not args.separator:
        raise SystemExit("--separator must not be empty")
    if not os.path.isfile(args.input):
        raise SystemExit(f"No such file or directory: '{args.input}'")

    data = convert_workbook(
        in_path=args.input,
        separator=args.separator,
        layer_names_sheet=args.layer_names_sheet,
        recenter_rules_sheet=args.recenter_rules_sheet,
        root_key=args.root_key,
    )

    if args.dry_run:
        body = data.get(args.root_key, {})
        layer_count = len([k for k in body.keys() if k != "recenterRules"])
        rule_count = len(cast(Dict[str, object], body.get("recenterRules", {})))
        print(
            f"Parsed {layer_count} layer-name keys and {rule_count} recenter rule groups"
        )
        return

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    indent = None if args.indent <= 0 else args.indent
    with open(args.output, "w", encoding="utf-8", newline="\n") as f:
        json.dump(data, f, ensure_ascii=False, indent=indent)
        f.write("\n")


if __name__ == "__main__":
    main()
