#!/usr/bin/env python3
"""
XLSX -> LAYER_NAME_CONFIG JSON converter.

Expected workbook shape:
- Sheet "LAYER_NAME_CONFIG_items" with columns: key, exact, contains
- Sheet "LAYER_NAME_CONFIG_recenterRules" with columns: force, noRecenter, alignH, alignV
- Sheet "TIMING_BEHAVIOR" (optional) with columns: layerName, behavior
- Sheet "TIMING_ITEM_SELECTOR" (optional) with columns: itemName, mode, value
- Sheet "SKIP_COPY_CONFIG" (optional) with columns: key, value, names

Rules:
- Sheet names are matched case-insensitively.
- exact/contains are split only by the explicit --separator argument.
- exact/contains are always emitted as arrays, even when empty.
"""

from __future__ import annotations

import argparse
import json
import os
from typing import TYPE_CHECKING, Dict, Iterable, List, Optional, Sequence, cast

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet as WorksheetType

try:
    from openpyxl import load_workbook as _openpyxl_load_workbook
except Exception:
    _openpyxl_load_workbook = None

RE_CENTER_KEYS: Sequence[str] = ("force", "noRecenter", "alignH", "alignV")
VALID_BEHAVIORS: Sequence[str] = ("timed", "span", "asIs")
VALID_SELECTOR_MODES: Sequence[str] = ("line", "index", "minMax")
SKIP_COPY_COMPLEX_KEYS: Sequence[str] = (
    "groups",
    "adHoc",
    "alwaysCopyLogoBaseNames",
)


def _json_scalar(value: object) -> str:
    return json.dumps(value, ensure_ascii=False)


def _format_indent_three(value: object, level: int = 0) -> str:
    if isinstance(value, dict):
        if not value:
            return "{}"
        indent = " " * 3 * level
        child_indent = " " * 3 * (level + 1)
        parts = []
        for key, item in value.items():
            rendered = _format_indent_three(item, level + 1)
            parts.append(f"{child_indent}{_json_scalar(key)}: {rendered}")
        return "{\n" + ",\n".join(parts) + "\n" + indent + "}"

    if isinstance(value, list):
        if not value:
            return "[]"
        return "[" + ", ".join(_json_scalar(item) for item in value) + "]"

    return _json_scalar(value)


def _norm_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def _sheet_by_name_ci(workbook: object, configured_name: str) -> "WorksheetType":
    target = configured_name.strip().lower()
    sheets = getattr(workbook, "worksheets", [])
    for ws in sheets:
        if str(getattr(ws, "title", "")).strip().lower() == target:
            return ws
    raise ValueError(f"Sheet not found (case-insensitive): {configured_name}")


def _sheet_by_name_ci_or_none(
    workbook: object, configured_name: str
) -> Optional["WorksheetType"]:
    target = configured_name.strip().lower()
    sheets = getattr(workbook, "worksheets", [])
    for ws in sheets:
        if str(getattr(ws, "title", "")).strip().lower() == target:
            return ws
    return None


def _split_list_cell(value: object, separator: str) -> List[str]:
    token = str(value or "").strip()
    if not token:
        return []
    return [part.strip() for part in token.split(separator) if part.strip()]


def _read_headers(worksheet: "WorksheetType") -> List[str]:
    row_iter = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
    first = next(row_iter, None)
    if first is None:
        raise ValueError("Worksheet is empty (missing header row)")
    return [str(c or "").strip() for c in first]


def _index_map(headers: Iterable[str]) -> Dict[str, int]:
    return {_norm_header(name): idx for idx, name in enumerate(headers)}


def _cell(row: Sequence[object], idx: int) -> object:
    return row[idx] if idx < len(row) else ""


def _parse_layer_names(
    worksheet: "WorksheetType", separator: str
) -> Dict[str, Dict[str, List[str]]]:
    headers = _read_headers(worksheet)
    idx = _index_map(headers)

    for required in ("key", "exact", "contains"):
        if required not in idx:
            raise ValueError(
                f"LAYER_NAME_CONFIG_items sheet is missing required column: {required}"
            )

    out: Dict[str, Dict[str, List[str]]] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        key = str(_cell(row, idx["key"]) or "").strip()
        if not key:
            continue

        exact = _split_list_cell(_cell(row, idx["exact"]), separator)
        contains = _split_list_cell(_cell(row, idx["contains"]), separator)
        out[key] = {"exact": exact, "contains": contains}

    return out


def _parse_recenter_rules(worksheet: "WorksheetType") -> Dict[str, List[str]]:
    headers = _read_headers(worksheet)
    idx = _index_map(headers)

    for required in RE_CENTER_KEYS:
        if _norm_header(required) not in idx:
            raise ValueError(
                f"LAYER_NAME_CONFIG_recenterRules sheet is missing required column: {required}"
            )

    out: Dict[str, List[str]] = {k: [] for k in RE_CENTER_KEYS}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for key in RE_CENTER_KEYS:
            col_idx = idx[_norm_header(key)]
            value = str(_cell(row, col_idx) or "").strip()
            if value:
                out[key].append(value)

    return out


def _parse_timing_behavior(worksheet: "WorksheetType") -> Dict[str, str]:
    headers = _read_headers(worksheet)
    idx = _index_map(headers)

    for required in ("layername", "behavior"):
        if required not in idx:
            raise ValueError(
                f"TIMING_BEHAVIOR sheet is missing required column: {required}"
            )

    out: Dict[str, str] = {}
    allowed = set(VALID_BEHAVIORS)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        layer_name = str(_cell(row, idx["layername"]) or "").strip()
        if not layer_name:
            continue
        behavior = str(_cell(row, idx["behavior"]) or "").strip()
        if not behavior:
            continue
        if behavior not in allowed:
            allowed_text = ", ".join(VALID_BEHAVIORS)
            raise ValueError(
                f"Invalid behavior '{behavior}' for layer '{layer_name}'. "
                f"Allowed values: {allowed_text}"
            )
        out[layer_name] = behavior

    return out


def _parse_timing_item_selector(
    worksheet: "WorksheetType",
) -> Dict[str, Dict[str, object]]:
    headers = _read_headers(worksheet)
    idx = _index_map(headers)

    for required in ("itemname", "mode", "value"):
        if required not in idx:
            raise ValueError(
                f"TIMING_ITEM_SELECTOR sheet is missing required column: {required}"
            )

    out: Dict[str, Dict[str, object]] = {}
    allowed = set(VALID_SELECTOR_MODES)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        item_name = str(_cell(row, idx["itemname"]) or "").strip()
        if not item_name:
            continue
        mode = str(_cell(row, idx["mode"]) or "").strip()
        if not mode:
            continue
        if mode not in allowed:
            allowed_text = ", ".join(VALID_SELECTOR_MODES)
            raise ValueError(
                f"Invalid mode '{mode}' for item '{item_name}'. "
                f"Allowed values: {allowed_text}"
            )

        raw_value = _cell(row, idx["value"])
        value: object
        if isinstance(raw_value, str):
            value = raw_value.strip()
        else:
            value = "" if raw_value is None else raw_value

        out[item_name] = {"mode": mode, "value": value}

    return out


def _parse_bool_cell(raw_value: object, *, key: str) -> bool:
    if isinstance(raw_value, bool):
        return raw_value
    if isinstance(raw_value, str):
        token = raw_value.strip().lower()
        if token == "true":
            return True
        if token == "false":
            return False
    raise ValueError(
        f"Invalid boolean value '{raw_value}' for key '{key}'. "
        "Allowed values: true, false"
    )


def _parse_skip_copy_config(
    worksheet: "WorksheetType", separator: str
) -> Dict[str, object]:
    headers = _read_headers(worksheet)
    idx = _index_map(headers)

    for required in ("key", "value", "names"):
        if required not in idx:
            raise ValueError(
                f"SKIP_COPY_CONFIG sheet is missing required column: {required}"
            )

    out: Dict[str, object] = {}
    complex_keys = set(SKIP_COPY_COMPLEX_KEYS)
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        key = str(_cell(row, idx["key"]) or "").strip()
        if not key:
            continue

        enabled = _parse_bool_cell(_cell(row, idx["value"]), key=key)
        names = _split_list_cell(_cell(row, idx["names"]), separator)

        if key in complex_keys:
            out[key] = {"enabled": enabled, "names": names}
        else:
            out[key] = enabled

    return out


def convert_workbook(
    in_path: str,
    separator: str,
    layer_names_sheet: str,
    recenter_rules_sheet: str,
    root_key: str,
    timing_behavior_sheet: Optional[str] = "TIMING_BEHAVIOR",
    timing_item_selector_sheet: Optional[str] = "TIMING_ITEM_SELECTOR",
    skip_config_sheet: Optional[str] = "SKIP_COPY_CONFIG",
) -> Dict[str, object]:
    if _openpyxl_load_workbook is None:
        raise RuntimeError(
            "XLSX support requires openpyxl. Install with: pip install openpyxl"
        )

    wb = _openpyxl_load_workbook(in_path, read_only=True, data_only=True)
    try:
        ws_layers = _sheet_by_name_ci(wb, layer_names_sheet)
        ws_rules = _sheet_by_name_ci(wb, recenter_rules_sheet)

        layer_map = _parse_layer_names(ws_layers, separator)
        rules_map = _parse_recenter_rules(ws_rules)

        body: Dict[str, object] = dict(layer_map)
        body["recenterRules"] = rules_map
        add_layers: Dict[str, object] = {root_key: body}
        if timing_behavior_sheet:
            ws_timing = _sheet_by_name_ci_or_none(wb, timing_behavior_sheet)
            if ws_timing is not None:
                add_layers["TIMING_BEHAVIOR"] = _parse_timing_behavior(ws_timing)
        if timing_item_selector_sheet:
            ws_selector = _sheet_by_name_ci_or_none(wb, timing_item_selector_sheet)
            if ws_selector is not None:
                add_layers["TIMING_ITEM_SELECTOR"] = _parse_timing_item_selector(
                    ws_selector
                )
        if skip_config_sheet:
            ws_skip = _sheet_by_name_ci_or_none(wb, skip_config_sheet)
            if ws_skip is not None:
                add_layers["SKIP_COPY_CONFIG"] = _parse_skip_copy_config(
                    ws_skip, separator
                )
        return {"config": {"addLayers": add_layers}}
    finally:
        wb.close()


def _write_json_output(path: str, data: object, indent: int) -> None:
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        if indent == 3:
            f.write(_format_indent_three(data))
        else:
            json.dump(
                data, f, ensure_ascii=False, indent=None if indent <= 0 else indent
            )
        f.write("\n")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert LAYER_NAME_CONFIG Excel workbook to JSON"
    )
    parser.add_argument("input", help="Path to input XLSX")
    parser.add_argument("output", help="Path to output JSON")
    parser.add_argument(
        "--separator",
        default=";",
        help="Explicit separator for exact/contains cells (default ';')",
    )
    parser.add_argument(
        "--layer-names-sheet",
        default="LAYER_NAME_CONFIG_items",
        help="Layer names sheet name (case-insensitive match, default LAYER_NAME_CONFIG_items)",
    )
    parser.add_argument(
        "--recenter-rules-sheet",
        default="LAYER_NAME_CONFIG_recenterRules",
        help="Recenter rules sheet name (case-insensitive match, default LAYER_NAME_CONFIG_recenterRules)",
    )
    parser.add_argument(
        "--timing-behavior-sheet",
        default="TIMING_BEHAVIOR",
        help="TIMING_BEHAVIOR sheet name (parsed by default when present; override to use a different sheet name)",
    )
    parser.add_argument(
        "--timing-item-selector-sheet",
        default="TIMING_ITEM_SELECTOR",
        help="TIMING_ITEM_SELECTOR sheet name (parsed by default when present; override to use a different sheet name)",
    )
    parser.add_argument(
        "--skip-config-sheet",
        default="SKIP_COPY_CONFIG",
        help="SKIP_COPY_CONFIG sheet name (parsed by default when present; override to use a different sheet name)",
    )
    parser.add_argument(
        "--root-key",
        default="LAYER_NAME_CONFIG",
        help="Root key in output JSON (default LAYER_NAME_CONFIG)",
    )
    parser.add_argument(
        "--indent",
        type=int,
        default=4,
        help="JSON indentation (default 4; set 0 for compact; set 3 for inline arrays)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and print summary only (no output file)",
    )
    args = parser.parse_args()

    if not args.separator:
        raise SystemExit("--separator must not be empty")
    if not os.path.isfile(args.input):
        raise SystemExit(f"No such file or directory: '{args.input}'")

    data = convert_workbook(
        in_path=args.input,
        separator=args.separator,
        layer_names_sheet=args.layer_names_sheet,
        recenter_rules_sheet=args.recenter_rules_sheet,
        root_key=args.root_key,
        timing_behavior_sheet=args.timing_behavior_sheet,
        timing_item_selector_sheet=args.timing_item_selector_sheet,
        skip_config_sheet=args.skip_config_sheet,
    )

    if args.dry_run:
        config = cast(Dict[str, object], data.get("config", {}))
        add_layers = cast(Dict[str, object], config.get("addLayers", {}))
        body = cast(Dict[str, object], add_layers.get(args.root_key, {}))
        layer_count = len([k for k in body.keys() if k != "recenterRules"])
        rule_count = len(cast(Dict[str, object], body.get("recenterRules", {})))
        extra = ""
        if "TIMING_BEHAVIOR" in add_layers:
            timing_count = len(cast(Dict[str, object], add_layers["TIMING_BEHAVIOR"]))
            extra = f"; {timing_count} TIMING_BEHAVIOR entries"
        if "TIMING_ITEM_SELECTOR" in add_layers:
            selector_count = len(
                cast(Dict[str, object], add_layers["TIMING_ITEM_SELECTOR"])
            )
            extra += f"; {selector_count} TIMING_ITEM_SELECTOR entries"
        if "SKIP_COPY_CONFIG" in add_layers:
            skip_count = len(cast(Dict[str, object], add_layers["SKIP_COPY_CONFIG"]))
            extra += f"; {skip_count} SKIP_COPY_CONFIG entries"
        print(
            f"Parsed {layer_count} layer-name keys and {rule_count} recenter rule groups{extra}"
        )
        return

    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)
    _write_json_output(args.output, data, args.indent)


if __name__ == "__main__":
    main()
