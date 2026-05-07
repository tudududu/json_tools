from __future__ import annotations

import csv
import os
from typing import TYPE_CHECKING, Callable, List, Optional, Type

from .forward import HEADER

try:
    from ..xlsx_styling import (
        apply_table_style,
        autosize_columns,
        read_theme_xml_bytes as _read_theme_xml_bytes,
        resolve_optional_file_path as _resolve_optional_file_path,
    )
except ImportError:
    from python.tools.xlsx_styling import (  # type: ignore[no-redef]
        apply_table_style,
        autosize_columns,
        read_theme_xml_bytes as _read_theme_xml_bytes,
        resolve_optional_file_path as _resolve_optional_file_path,
    )

# openpyxl is an optional runtime dependency (only needed for XLSX output).
# The TYPE_CHECKING block gives Pylance accurate type information without
# requiring the package to be installed. The try/except block captures the
# actual runtime values; all three names are checked together before use.
if TYPE_CHECKING:
    from openpyxl import Workbook as WorkbookType
    from openpyxl.styles import Color as ColorType
    from openpyxl.styles import Font as FontType
    from openpyxl.styles import PatternFill as PatternFillType

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Color, Font, PatternFill
except Exception:  # pragma: no cover - optional dependency
    Workbook = None  # type: ignore[assignment,misc]
    load_workbook = None  # type: ignore[assignment,misc]
    Color = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]

_Workbook: Optional[Type["WorkbookType"]] = Workbook
_load_workbook: Optional[Callable[..., "WorkbookType"]] = load_workbook
_Color: Optional[Type["ColorType"]] = Color
_Font: Optional[Type["FontType"]] = Font
_PatternFill: Optional[Type["PatternFillType"]] = PatternFill

XLSX_TEMPLATE_ENV = "SRT_TO_CSV_XLSX_TEMPLATE"
XLSX_THEME_ENV = "SRT_TO_CSV_XLSX_THEME_FILE"
DEFAULT_XLSX_THEME_FILE = os.path.normpath(
    os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..",
        "themes",
        "subtitles_theme.xml",
    )
)

XLSX_HEADER = HEADER + [f"<ISO>{i}" for i in range(1, 11)]


def _create_output_workbook(
    template_path_raw: Optional[str],
    template_source: str,
) -> "WorkbookType":
    """Create workbook, optionally from template, with explicit error handling.

    If a template is configured, loading must succeed; otherwise we
    abort with a clear message instead of silently falling back to default
    theme/workbook styling.
    """
    if _Workbook is None:
        raise SystemExit(
            "XLSX output requires openpyxl. Install with: pip install openpyxl"
        )

    template_path = _resolve_optional_file_path(template_path_raw)
    if not template_path:
        return _Workbook()

    if _load_workbook is None:
        raise SystemExit(
            "XLSX template loading requires openpyxl load_workbook support"
        )
    if not os.path.isfile(template_path):
        raise SystemExit(
            f"XLSX template path from {template_source} was not found: {template_path}"
        )

    try:
        return _load_workbook(template_path)
    except Exception as ex:
        raise SystemExit(
            f"Failed to load XLSX template from {template_source}: {template_path} ({ex})"
        )


def _resolve_theme_xml_bytes(
    template_path_raw: Optional[str],
    theme_path_raw: Optional[str],
) -> Optional[bytes]:
    """Resolve theme bytes for output workbook.

    Priority:
    1) --xlsx-theme-file
    2) SRT_TO_CSV_XLSX_THEME_FILE
    3) Built-in default theme file (only when no template workbook is used)
    """
    if theme_path_raw:
        return _read_theme_xml_bytes(theme_path_raw, "--xlsx-theme-file")

    env_theme_path = os.getenv(XLSX_THEME_ENV)
    if env_theme_path and env_theme_path.strip():
        return _read_theme_xml_bytes(env_theme_path, XLSX_THEME_ENV)

    # Preserve workbook theme from a provided template unless explicitly overridden.
    if _resolve_optional_file_path(template_path_raw):
        return None

    if os.path.isfile(DEFAULT_XLSX_THEME_FILE):
        return _read_theme_xml_bytes(DEFAULT_XLSX_THEME_FILE, "built-in theme")

    return None


def write_tabular_output(
    out_path: str,
    rows: List[List[str]],
    quote_all: bool,
    delimiter_name: str,
    output_type: str,
    xlsx_template: Optional[str] = None,
    xlsx_theme_file: Optional[str] = None,
) -> None:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    if output_type == "csv":
        delimiter = "," if delimiter_name == "comma" else ";"
        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(
                f,
                delimiter=delimiter,
                quoting=(csv.QUOTE_ALL if quote_all else csv.QUOTE_MINIMAL),
            )
            writer.writerow(HEADER)
            writer.writerows(rows)
        return

    if output_type != "xlsx":
        raise ValueError("output_type must be 'csv' or 'xlsx'")
    # openpyxl symbols are checked together because XLSX formatting relies on
    # workbook, styles, and table objects.
    # Checking all needed symbols keeps the type narrowed to
    # non-None for the remainder of the XLSX path.
    if (
        _Workbook is None
        or _Color is None
        or _Font is None
        or _PatternFill is None
    ):
        raise SystemExit(
            "XLSX output requires openpyxl. Install with: pip install openpyxl"
        )

    resolved_template = xlsx_template
    template_source = "--xlsx-template"
    if not _resolve_optional_file_path(resolved_template):
        resolved_template = os.getenv(XLSX_TEMPLATE_ENV)
        template_source = XLSX_TEMPLATE_ENV
    wb = _create_output_workbook(resolved_template, template_source)
    theme_xml = _resolve_theme_xml_bytes(resolved_template, xlsx_theme_file)
    if theme_xml is not None:
        wb.loaded_theme = theme_xml

    ws = wb.active
    if ws is None:
        raise SystemExit("Failed to create XLSX worksheet")

    # Ensure the output worksheet starts from a clean slate while preserving
    # any workbook-level theme inherited from a template workbook.
    ws.delete_rows(1, ws.max_row)
    ws.tables.clear()
    ws.title = "subtitles"
    ws.append(XLSX_HEADER)

    # Requested widths in openpyxl column units; reproduce original fixed schema.
    autosize_columns(
        ws,
        manual_width_overrides={
            "A": 12,
            "B": 12,
            "C": 41,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 16,
            "H": 16,
            "I": 16,
            "J": 16,
            "K": 16,
            "L": 16,
            "M": 16,
        },
    )

    # Excel theme color: Plum, Accent 5, Lighter 80%.
    title_fill = _PatternFill(fill_type="solid", fgColor=_Color(theme=8, tint=0.8))
    body_font = _Font(name="Aptos Narrow", size=12)

    for row in rows:
        normalized = list(row[: len(XLSX_HEADER)])
        if len(normalized) < len(XLSX_HEADER):
            normalized.extend([""] * (len(XLSX_HEADER) - len(normalized)))
        ws.append(normalized)

    # Apply body font to data rows only; keep header typography controlled by
    # the table style (for example white + bold in Medium styles).
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(XLSX_HEADER) + 1):
            ws.cell(row=row_idx, column=col_idx).font = body_font

    # Format the data range as a table with headers.
    if ws.max_row >= 2:
        apply_table_style(ws, table_name="SubtitlesTable")

    # Re-apply joined title row fill after table insertion so it remains visible.
    for row_idx in range(2, ws.max_row + 1):
        c1 = ws.cell(row=row_idx, column=1).value
        c2 = ws.cell(row=row_idx, column=2).value
        c3 = ws.cell(row=row_idx, column=3).value

        # Only style rows that contain at least one value in the data columns.
        in_use = any(value not in (None, "") for value in (c1, c2, c3))
        if not in_use:
            continue

        # Joined-output title marker rows are emitted as ["", "", <filename>].
        is_title_row = (
            c1 in (None, "") and c2 in (None, "") and str(c3 or "").strip() != ""
        )
        if is_title_row:
            for col_idx in range(1, len(XLSX_HEADER) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = title_fill

    wb.save(out_path)
