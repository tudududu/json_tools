from __future__ import annotations

import csv
import os
import re
from typing import List, Optional, Tuple

from python.tools.srt_csv.timecode import FRAME_TC_RE, MS_TC_RE, format_time_ms

ISO_HEADER_RE = re.compile(r"^[A-Z]{3}(?:_[A-Z]{3})?$")

try:
    from openpyxl import load_workbook as _load_workbook
except Exception:  # pragma: no cover - optional dependency
    _load_workbook = None  # type: ignore[assignment]


def _normalize_header_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (name or "").strip().lower())


def _normalize_iso_header(name: str) -> Optional[str]:
    token = (name or "").strip().upper()
    if ISO_HEADER_RE.fullmatch(token):
        return token
    return None


def _resolve_column_index(
    headers: List[str],
    override: Optional[str],
    aliases: Tuple[str, ...],
) -> int:
    if override:
        if override.isdigit():
            idx = int(override) - 1
            if 0 <= idx < len(headers):
                return idx
            raise ValueError(f"Column index out of range: {override}")
        target = _normalize_header_name(override)
        for i, header in enumerate(headers):
            if _normalize_header_name(header) == target:
                return i
        raise ValueError(f"Column not found: {override}")

    normalized_headers = [_normalize_header_name(h) for h in headers]
    for alias in aliases:
        alias_norm = _normalize_header_name(alias)
        if alias_norm in normalized_headers:
            return normalized_headers.index(alias_norm)
    raise ValueError(f"Could not detect required column. Tried aliases: {aliases}")


def _resolve_iso_text_columns(
    headers: List[str],
    text_col_filter: Optional[str],
) -> List[Tuple[int, str]]:
    iso_columns: List[Tuple[int, str]] = []
    for i, header in enumerate(headers):
        iso = _normalize_iso_header(header)
        if iso:
            iso_columns.append((i, iso))

    if not iso_columns:
        return []

    if not text_col_filter:
        return iso_columns

    if text_col_filter.isdigit():
        idx = int(text_col_filter) - 1
        if idx < 0 or idx >= len(headers):
            raise ValueError(f"Column index out of range: {text_col_filter}")
        return [(i, iso) for i, iso in iso_columns if i == idx]

    target = _normalize_header_name(text_col_filter)
    return [
        (i, iso)
        for i, iso in iso_columns
        if _normalize_header_name(headers[i]) == target
    ]


def _read_reverse_table(
    in_path: str,
    encoding: str,
) -> Tuple[List[str], List[List[str]]]:
    if in_path.lower().endswith(".xlsx"):
        if _load_workbook is None:
            raise SystemExit(
                "XLSX input requires openpyxl. Install with: pip install openpyxl"
            )
        wb = _load_workbook(in_path, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("XLSX file has no active worksheet")
        all_rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(["" if v is None else str(v) for v in row])
        if not all_rows:
            return [], []
        headers = [c.strip() for c in all_rows[0]]
        return headers, all_rows[1:]

    with open(in_path, "r", newline="", encoding=encoding) as f:
        sample = f.read(8192)
        f.seek(0)
        delimiter = ","
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;")
            delimiter = dialect.delimiter
        except Exception:
            delimiter = ";" if ";" in sample and "," not in sample else ","
        reader = csv.reader(f, delimiter=delimiter)
        all_rows = [list(r) for r in reader]
    if not all_rows:
        return [], []
    headers = [(c or "").strip() for c in all_rows[0]]
    return headers, all_rows[1:]


def _detect_reverse_time_format(
    rows: List[List[str]],
    idx_start: int,
    idx_end: int,
) -> str:
    detected: Optional[str] = None
    for row in rows:
        start = (row[idx_start] if idx_start < len(row) else "").strip()
        end = (row[idx_end] if idx_end < len(row) else "").strip()
        if not start and not end:
            continue
        if not start or not end:
            raise ValueError(
                "Row has only one time value; both Start Time and End Time are required"
            )
        for token in (start, end):
            current: Optional[str] = None
            if FRAME_TC_RE.fullmatch(token):
                current = "frames"
            elif MS_TC_RE.fullmatch(token):
                current = "ms"
            else:
                raise ValueError(f"Unsupported timecode format: {token}")
            if detected is None:
                detected = current
            elif detected != current:
                raise ValueError("Mixed timecode formats detected in a single file")
    if detected is None:
        raise ValueError("No valid timed subtitle rows found in input")
    return detected


def _parse_reverse_timecode(value: str, fmt: str, fps: float) -> float:
    token = value.strip()
    if fmt == "frames":
        h, m, s, ff = [int(part) for part in token.split(":")]
        if fps <= 0:
            raise ValueError("fps must be > 0 for frames input")
        return h * 3600 + m * 60 + s + ff / fps
    if fmt == "ms":
        base, ms = re.split(r"[,.]", token, maxsplit=1)
        h, m, s = [int(part) for part in base.split(":")]
        return h * 3600 + m * 60 + s + int(ms) / 1000.0
    raise ValueError(f"Unsupported timecode format: {fmt}")


def _rows_to_reverse_records(
    rows: List[List[str]],
    idx_start: int,
    idx_end: int,
    idx_text: int,
    time_format: str,
    fps: float,
) -> List[Tuple[float, float, str]]:
    out: List[Tuple[float, float, str]] = []
    for row in rows:
        start = (row[idx_start] if idx_start < len(row) else "").strip()
        end = (row[idx_end] if idx_end < len(row) else "").strip()
        text = row[idx_text] if idx_text < len(row) else ""

        # Skip joined marker rows emitted by --join-output in forward mode.
        if not start and not end:
            continue
        if not start or not end:
            raise ValueError(
                "Row has only one time value; both Start Time and End Time are required"
            )

        tin = _parse_reverse_timecode(start, time_format, fps)
        tout = _parse_reverse_timecode(end, time_format, fps)
        out.append((tin, tout, text))
    return out


def _records_to_srt_text(records: List[Tuple[float, float, str]]) -> str:
    lines: List[str] = []
    for idx, (start, end, text) in enumerate(records, start=1):
        lines.append(str(idx))
        lines.append(f"{format_time_ms(start)} --> {format_time_ms(end)}")
        text_lines = text.splitlines() if text else [""]
        lines.extend(text_lines)
        lines.append("")
    return "\n".join(lines)


def _extract_joined_reverse_blocks(
    rows: List[List[str]],
    idx_start: int,
    idx_end: int,
    idx_text: int,
) -> List[Tuple[str, List[List[str]]]]:
    blocks: List[Tuple[str, List[List[str]]]] = []
    current_name: Optional[str] = None
    current_rows: List[List[str]] = []
    marker_count = 0

    for row in rows:
        start = (row[idx_start] if idx_start < len(row) else "").strip()
        end = (row[idx_end] if idx_end < len(row) else "").strip()
        text = (row[idx_text] if idx_text < len(row) else "").strip()

        if not start and not end:
            if not text:
                continue
            marker_count += 1
            if current_name is not None:
                blocks.append((current_name, current_rows))
            current_name = text
            current_rows = []
            continue

        if not start or not end:
            raise ValueError(
                "Row has only one time value; both Start Time and End Time are required"
            )
        if current_name is None:
            raise ValueError(
                "Joined reverse input requires marker rows before timed rows"
            )
        current_rows.append(row)

    if current_name is not None:
        blocks.append((current_name, current_rows))

    if marker_count == 0:
        raise ValueError(
            "Joined reverse input requires marker rows (empty Start/End with filename in Text)"
        )

    return blocks


def _sanitize_joined_marker_filename(marker: str) -> str:
    raw = (marker or "").strip()
    raw = re.sub(r"\.srt$", "", raw, flags=re.I)
    raw = re.sub(r"[\\/]+", "_", raw)
    raw = re.sub(r"[^A-Za-z0-9._ -]", "_", raw)
    raw = re.sub(r"\s+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("._ ")
    if not raw:
        raw = "subtitle"
    return f"{raw}.srt"


def _dedupe_output_filename(name: str, used: set[str]) -> str:
    base, ext = os.path.splitext(name)
    candidate = name
    i = 2
    while candidate.lower() in used:
        candidate = f"{base}_{i}{ext}"
        i += 1
    used.add(candidate.lower())
    return candidate


def _prepare_used_filenames(out_dir: str) -> set[str]:
    used: set[str] = set()
    if not os.path.isdir(out_dir):
        return used
    for name in os.listdir(out_dir):
        if os.path.isfile(os.path.join(out_dir, name)):
            used.add(name.lower())
    return used


def csv_to_srt(
    in_path: str,
    out_path: str,
    fps: float,
    encoding: str,
    start_col: Optional[str] = None,
    end_col: Optional[str] = None,
    text_col: Optional[str] = None,
) -> None:
    headers, rows = _read_reverse_table(in_path, encoding=encoding)
    if not headers:
        raise ValueError(f"Input table is empty: {in_path}")

    idx_start = _resolve_column_index(
        headers,
        start_col,
        aliases=("Start Time", "start", "in", "inpoint"),
    )
    idx_end = _resolve_column_index(
        headers,
        end_col,
        aliases=("End Time", "end", "out", "outpoint"),
    )

    country_columns = _resolve_iso_text_columns(headers, text_col)
    time_format = _detect_reverse_time_format(rows, idx_start, idx_end)

    out_dir = os.path.dirname(out_path) or "."
    out_base = os.path.splitext(os.path.basename(out_path))[0]
    os.makedirs(out_dir, exist_ok=True)

    # Multi-country mode when ISO headers are present (or filtered to ISO headers).
    if country_columns:
        used_names = _prepare_used_filenames(out_dir)
        for idx_text, iso in country_columns:
            records = _rows_to_reverse_records(
                rows,
                idx_start=idx_start,
                idx_end=idx_end,
                idx_text=idx_text,
                time_format=time_format,
                fps=fps,
            )
            if not records:
                continue
            fname = _dedupe_output_filename(f"{out_base}_{iso}.srt", used_names)
            out_country_path = os.path.join(out_dir, fname)
            with open(out_country_path, "w", encoding="utf-8", newline="\n") as f:
                f.write(_records_to_srt_text(records))
        return

    # Fallback to legacy single text-column behavior.
    idx_text = _resolve_column_index(
        headers,
        text_col,
        aliases=("Text", "subtitle", "caption"),
    )
    records = _rows_to_reverse_records(
        rows,
        idx_start=idx_start,
        idx_end=idx_end,
        idx_text=idx_text,
        time_format=time_format,
        fps=fps,
    )

    with open(out_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(_records_to_srt_text(records))


def csv_to_srt_joined(
    in_path: str,
    out_dir: str,
    fps: float,
    encoding: str,
    start_col: Optional[str] = None,
    end_col: Optional[str] = None,
    text_col: Optional[str] = None,
) -> List[str]:
    headers, rows = _read_reverse_table(in_path, encoding=encoding)
    if not headers:
        raise ValueError(f"Input table is empty: {in_path}")

    idx_start = _resolve_column_index(
        headers,
        start_col,
        aliases=("Start Time", "start", "in", "inpoint"),
    )
    idx_end = _resolve_column_index(
        headers,
        end_col,
        aliases=("End Time", "end", "out", "outpoint"),
    )

    country_columns = _resolve_iso_text_columns(headers, text_col)

    os.makedirs(out_dir, exist_ok=True)
    used_names: set[str] = _prepare_used_filenames(out_dir)
    written: List[str] = []

    if country_columns:
        marker_idx = None
        try:
            marker_idx = _resolve_column_index(
                headers,
                None,
                aliases=("Text", "subtitle", "caption"),
            )
        except ValueError:
            marker_idx = country_columns[0][0]

        blocks = _extract_joined_reverse_blocks(rows, idx_start, idx_end, marker_idx)
        timed_rows = [r for _, block_rows in blocks for r in block_rows]
        time_format = _detect_reverse_time_format(timed_rows, idx_start, idx_end)

        for marker_name, block_rows in blocks:
            if not block_rows:
                continue
            marker_base = os.path.splitext(
                _sanitize_joined_marker_filename(marker_name)
            )[0]
            for idx_text, iso in country_columns:
                records = _rows_to_reverse_records(
                    block_rows,
                    idx_start=idx_start,
                    idx_end=idx_end,
                    idx_text=idx_text,
                    time_format=time_format,
                    fps=fps,
                )
                if not records:
                    continue
                fname = _dedupe_output_filename(f"{marker_base}_{iso}.srt", used_names)
                out_path = os.path.join(out_dir, fname)
                with open(out_path, "w", encoding="utf-8", newline="\n") as f:
                    f.write(_records_to_srt_text(records))
                written.append(out_path)

        if not written:
            raise ValueError("No timed subtitle blocks found under joined markers")
        return written

    idx_text = _resolve_column_index(
        headers,
        text_col,
        aliases=("Text", "subtitle", "caption"),
    )
    blocks = _extract_joined_reverse_blocks(rows, idx_start, idx_end, idx_text)
    timed_rows = [r for _, block_rows in blocks for r in block_rows]
    time_format = _detect_reverse_time_format(timed_rows, idx_start, idx_end)

    for marker_name, block_rows in blocks:
        if not block_rows:
            continue
        records = _rows_to_reverse_records(
            block_rows,
            idx_start=idx_start,
            idx_end=idx_end,
            idx_text=idx_text,
            time_format=time_format,
            fps=fps,
        )
        if not records:
            continue
        fname = _sanitize_joined_marker_filename(marker_name)
        fname = _dedupe_output_filename(fname, used_names)
        out_path = os.path.join(out_dir, fname)
        with open(out_path, "w", encoding="utf-8", newline="\n") as f:
            f.write(_records_to_srt_text(records))
        written.append(out_path)

    if not written:
        raise ValueError("No timed subtitle blocks found under joined markers")
    return written


__all__ = [
    "_normalize_header_name",
    "_normalize_iso_header",
    "_resolve_column_index",
    "_resolve_iso_text_columns",
    "_read_reverse_table",
    "_detect_reverse_time_format",
    "_parse_reverse_timecode",
    "_rows_to_reverse_records",
    "_records_to_srt_text",
    "_extract_joined_reverse_blocks",
    "_sanitize_joined_marker_filename",
    "_dedupe_output_filename",
    "csv_to_srt",
    "csv_to_srt_joined",
]
