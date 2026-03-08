import os
import tempfile
import unittest

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

from python import csv_to_json as mod


class ParseTimecodeTests(unittest.TestCase):
    def test_plain_seconds(self):
        self.assertAlmostEqual(mod.parse_timecode('12', 25), 12.0)
        self.assertAlmostEqual(mod.parse_timecode('12.5', 25), 12.5)
        self.assertAlmostEqual(mod.parse_timecode('12,5', 25), 12.5)

    def test_hh_mm_ss_ff(self):
        # 00:00:01:12 at 24 fps => 1.5s
        self.assertAlmostEqual(mod.parse_timecode('00:00:01:12', 24), 1.5)

    def test_hh_mm_ss_ms(self):
        self.assertAlmostEqual(mod.parse_timecode('00:01:02.5', 25), 62.5)

    def test_mm_ss(self):
        self.assertAlmostEqual(mod.parse_timecode('02:03.5', 25), 123.5)

    def test_invalid_inputs(self):
        with self.assertRaises(ValueError):
            mod.parse_timecode(None, 25)  # type: ignore[arg-type]
        with self.assertRaises(ValueError):
            mod.parse_timecode('', 25)
        with self.assertRaises(ValueError):
            mod.parse_timecode('bad:format', 25)


class SniffDelimiterTests(unittest.TestCase):
    def test_preferred_named_mapping(self):
        sample = 'a;b;c\n1;2;3\n'
        self.assertEqual(mod._sniff_delimiter(sample, preferred='semicolon'), ';')
        self.assertEqual(mod._sniff_delimiter(sample, preferred=';'), ';')

    def test_auto_detect(self):
        sample = 'a\tb\tc\n1\t2\t3\n'
        self.assertEqual(mod._sniff_delimiter(sample, preferred='auto'), '\t')
        # No delimiters at all -> fallback to comma
        self.assertEqual(mod._sniff_delimiter('abc\nxyz\n', preferred=None), ',')


class DetectColumnsTests(unittest.TestCase):
    def test_detect_basic(self):
        headers = ['Start Time', 'End', 'Caption']
        s, e, t = mod.detect_columns(headers)
        self.assertEqual((s, e, t), ('Start Time', 'End', 'Caption'))

    def test_overrides_by_index(self):
        headers = ['S', 'E', 'T']
        s, e, t = mod.detect_columns(headers, start_override='1', end_override='2', text_override='3')
        self.assertEqual((s, e, t), ('S', 'E', 'T'))


class SafeIntTests(unittest.TestCase):
    def test_safe_int(self):
        self.assertEqual(mod.safe_int('5'), 5)
        self.assertEqual(mod.safe_int('x', default=7), 7)


class ConvertMinimalTests(unittest.TestCase):
    def test_minimal_no_rows(self):
        # A simple-mode header with no rows returns empty subtitles list
        csv_content = 'Start Time,End Time,Text\n'
        with tempfile.NamedTemporaryFile('w+', delete=False, suffix='.csv') as f:
            f.write(csv_content)
            path = f.name
        try:
            out = mod.convert_csv_to_json(path)
            self.assertEqual(out, {"subtitles": []})
        finally:
            os.remove(path)


@unittest.skipUnless(Workbook is not None, "openpyxl is required for XLSX tests")
class ConvertXlsxTests(unittest.TestCase):
    def test_simple_xlsx_convert(self):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            path = f.name
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "data"
            ws.append(["Start Time", "End Time", "Text"])
            ws.append(["0", "1.2", "Hello XLSX"])
            wb.save(path)
            wb.close()

            out = mod.convert_csv_to_json(path)
            self.assertEqual(out["subtitles"][0]["text"], "Hello XLSX")
        finally:
            os.remove(path)

    def test_xlsx_sheet_selection_default_data_else_first(self):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as f:
            path = f.name
        try:
            wb = Workbook()
            ws_first = wb.active
            ws_first.title = "Sheet1"
            ws_first.append(["Start Time", "End Time", "Text"])
            ws_first.append(["0", "1", "Wrong sheet"])

            ws_data = wb.create_sheet("data")
            ws_data.append(["Start Time", "End Time", "Text"])
            ws_data.append(["0", "1", "Data sheet"])
            wb.save(path)
            wb.close()

            out_default = mod.convert_csv_to_json(path)
            self.assertEqual(out_default["subtitles"][0]["text"], "Data sheet")

            out_sheet1 = mod.convert_csv_to_json(path, xlsx_sheet="Sheet1")
            self.assertEqual(out_sheet1["subtitles"][0]["text"], "Wrong sheet")
        finally:
            os.remove(path)


if __name__ == '__main__':
    unittest.main(verbosity=2)
