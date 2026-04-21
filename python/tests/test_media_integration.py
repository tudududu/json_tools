import os
import json
import io
import tempfile
import unittest
from unittest import mock

try:
    from openpyxl import Workbook
except Exception:
    Workbook = None

from python import csv_to_json as mod


def tmp_file(suffix=".csv"):
    f = tempfile.NamedTemporaryFile("w+", delete=False, suffix=suffix)
    f.close()
    return f.name


class MediaIntegrationTests(unittest.TestCase):
    @staticmethod
    @unittest.skipUnless(Workbook is not None, "openpyxl is required for XLSX tests")
    def _write_layer_config_xlsx(path: str):
        wb = Workbook()
        ws_layers = wb.active
        ws_layers.title = "LAYER_NAME_CONFIG_items"
        ws_layers.append(["key", "exact", "contains"])
        ws_layers.append(["logo", "logo_01;Size_Holder_Logo", ""])
        ws_rules = wb.create_sheet(title="LAYER_NAME_CONFIG_recenterRules")
        ws_rules.append(["force", "noRecenter", "alignH", "alignV"])
        ws_rules.append(["Logo", "BG", "Claim", "Disclaimer"])
        ws_tb = wb.create_sheet(title="TIMING_BEHAVIOR")
        ws_tb.append(["layerName", "behavior"])
        ws_tb.append(["logo", "timed"])
        ws_tis = wb.create_sheet(title="TIMING_ITEM_SELECTOR")
        ws_tis.append(["itemName", "mode", "value"])
        ws_tis.append(["logo", "line", 1])
        ws_skip = wb.create_sheet(title="SKIP_COPY_CONFIG")
        ws_skip.append(["key", "value", "names"])
        ws_skip.append(["groups", "TRUE", "info; claim"])
        ws_skip.append(["disclaimerOff", "TRUE", ""])
        wb.save(path)
        wb.close()

    def test_media_injected_for_exact_country_language(self):
        # Unified CSV with two countries; language meta_global empty
        csv_content = (
            "record_type;video_id;line;start;end;key;is_global;country_scope;metadata;GBL;FRA;GBL;FRA\n"
            "meta_global;;;;;briefVersion;Y;ALL;6;;;;\n"
            "meta_global;;;;;fps;Y;ALL;25;;;;\n"
            "meta_global;;;;;language;Y;ALL;;;;;\n"
            "meta_local;V;;;;title;N;ALL;T;;;;\n"
            "sub;V;1;00:00:00:00;00:00:01:00;;;;;;;;x;y\n"
        )
        media_csv = (
            "AspectRatio;Dimensions;Creative;Media;Template;Template_name;Country;Language\n"
            "1x1;640x640;06sC1;TikTok;regular;;GBL;\n"
            "9x16;720x1280;15sC1;Meta InFeed;extra;tiktok;GBL;\n"
        )
        in_path = tmp_file(".csv")
        with open(in_path, "w", encoding="utf-8") as f:
            f.write(csv_content)
        media_path = tmp_file(".csv")
        with open(media_path, "w", encoding="utf-8") as f:
            f.write(media_csv)
        try:
            with tempfile.TemporaryDirectory() as td:
                pattern = os.path.join(td, "out-{country}.json")
                rc = mod.main(
                    [
                        in_path,
                        pattern,
                        "--split-by-country",
                        "--media-config",
                        media_path,
                    ]
                )
                self.assertEqual(rc, 0)
                gbl_path = os.path.join(td, "out-GBL.json")
                fra_path = os.path.join(td, "out-FRA.json")
                self.assertTrue(os.path.isfile(gbl_path))
                self.assertTrue(os.path.isfile(fra_path))
                with open(gbl_path, "r", encoding="utf-8") as f:
                    gbl = json.load(f)
                with open(fra_path, "r", encoding="utf-8") as f:
                    fra = json.load(f)
                self.assertIn("config", gbl)
                self.assertIn("pack", gbl["config"])
                self.assertIn("EXTRA_OUTPUT_COMPS", gbl["config"]["pack"])
                self.assertNotIn("config", fra)
                self.assertNotIn("media", gbl)
                self.assertNotIn("media", fra)
                # Basic sanity of media content
                self.assertIn("1x1|06s", gbl["config"]["pack"]["EXTRA_OUTPUT_COMPS"])
        finally:
            try:
                os.remove(in_path)
            except Exception:
                pass
            try:
                os.remove(media_path)
            except Exception:
                pass

    @unittest.skipUnless(Workbook is not None, "openpyxl is required for XLSX tests")
    def test_media_injected_from_xlsx_media_file(self):
        csv_content = (
            "record_type;video_id;line;start;end;key;is_global;country_scope;metadata;GBL;FRA;GBL;FRA\n"
            "meta_global;;;;;briefVersion;Y;ALL;6;;;;\n"
            "meta_global;;;;;fps;Y;ALL;25;;;;\n"
            "meta_global;;;;;language;Y;ALL;;;;;\n"
            "meta_local;V;;;;title;N;ALL;T;;;;\n"
            "sub;V;1;00:00:00:00;00:00:01:00;;;;;;;;x;y\n"
        )
        in_path = tmp_file(".csv")
        with open(in_path, "w", encoding="utf-8") as f:
            f.write(csv_content)

        media_path = tmp_file(".xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "media"
        ws.append(
            [
                "AspectRatio",
                "Dimensions",
                "Creative",
                "Media",
                "Template",
                "Template_name",
                "Country",
                "Language",
            ]
        )
        ws.append(["1x1", "640x640", "06sC1", "TikTok", "regular", "", "GBL", ""])
        wb.save(media_path)
        wb.close()

        try:
            with tempfile.TemporaryDirectory() as td:
                pattern = os.path.join(td, "out-{country}.json")
                rc = mod.main(
                    [
                        in_path,
                        pattern,
                        "--split-by-country",
                        "--media-config",
                        media_path,
                    ]
                )
                self.assertEqual(rc, 0)
                gbl_path = os.path.join(td, "out-GBL.json")
                fra_path = os.path.join(td, "out-FRA.json")
                self.assertTrue(os.path.isfile(gbl_path))
                self.assertTrue(os.path.isfile(fra_path))
                with open(gbl_path, "r", encoding="utf-8") as f:
                    gbl = json.load(f)
                with open(fra_path, "r", encoding="utf-8") as f:
                    fra = json.load(f)
                self.assertIn("config", gbl)
                self.assertIn("pack", gbl["config"])
                self.assertIn("EXTRA_OUTPUT_COMPS", gbl["config"]["pack"])
                self.assertNotIn("config", fra)
                self.assertNotIn("media", gbl)
                self.assertNotIn("media", fra)
                self.assertIn("1x1|06s", gbl["config"]["pack"]["EXTRA_OUTPUT_COMPS"])
        finally:
            try:
                os.remove(in_path)
            except Exception:
                pass
            try:
                os.remove(media_path)
            except Exception:
                pass

    @unittest.skipUnless(Workbook is not None, "openpyxl is required for XLSX tests")
    def test_layer_config_injected_under_config_add_layers(self):
        csv_content = (
            "record_type;video_id;line;start;end;key;is_global;country_scope;metadata;GBL\n"
            "meta_global;;;;;briefVersion;Y;ALL;6;\n"
            "meta_global;;;;;fps;Y;ALL;25;\n"
            "meta_global;;;;;language;Y;ALL;;\n"
            "meta_local;V;;;;title;N;ALL;T;\n"
            "sub;V;1;00:00:00:00;00:00:01:00;;;;;;;x\n"
        )
        in_path = tmp_file(".csv")
        with open(in_path, "w", encoding="utf-8") as f:
            f.write(csv_content)

        layer_cfg_path = tmp_file(".xlsx")
        self._write_layer_config_xlsx(layer_cfg_path)

        try:
            with tempfile.TemporaryDirectory() as td:
                out_json = os.path.join(td, "out.json")
                rc = mod.main([in_path, out_json, "--layer-config", layer_cfg_path])
                self.assertEqual(rc, 0)
                with open(out_json, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                self.assertIn("config", payload)
                self.assertIn("addLayers", payload["config"])
                add_layers = payload["config"]["addLayers"]
                self.assertIn("LAYER_NAME_CONFIG", add_layers)
                self.assertIn("TIMING_BEHAVIOR", add_layers)
                self.assertIn("TIMING_ITEM_SELECTOR", add_layers)
                self.assertIn("SKIP_COPY_CONFIG", add_layers)
                self.assertIn("logo", add_layers["LAYER_NAME_CONFIG"])
                self.assertEqual(add_layers["TIMING_BEHAVIOR"].get("logo"), "timed")
                self.assertEqual(
                    add_layers["TIMING_ITEM_SELECTOR"].get("logo"),
                    {"mode": "line", "value": 1},
                )
                self.assertEqual(
                    add_layers["SKIP_COPY_CONFIG"].get("groups"),
                    {"enabled": True, "names": ["info", "claim"]},
                )
        finally:
            try:
                os.remove(in_path)
            except Exception:
                pass
            try:
                os.remove(layer_cfg_path)
            except Exception:
                pass

    @unittest.skipUnless(Workbook is not None, "openpyxl is required for XLSX tests")
    def test_layer_config_replaces_existing_add_layers(self):
        csv_content = (
            "record_type;video_id;line;start;end;key;is_global;country_scope;metadata;GBL\n"
            "meta_global;;;;;briefVersion;Y;ALL;6;\n"
            "meta_global;;;;;fps;Y;ALL;25;\n"
            "meta_global;;;;;language;Y;ALL;;\n"
            "meta_local;V;;;;title;N;ALL;T;\n"
            "sub;V;1;00:00:00:00;00:00:01:00;;;;;;;x\n"
        )
        in_path = tmp_file(".csv")
        with open(in_path, "w", encoding="utf-8") as f:
            f.write(csv_content)

        layer_cfg_path = tmp_file(".xlsx")
        self._write_layer_config_xlsx(layer_cfg_path)

        mocked_data = {
            "metadataGlobal": {"briefVersion": 6, "fps": 25, "language": ""},
            "videos": [
                {
                    "videoId": "V_landscape",
                    "metadata": {"orientation": "landscape"},
                    "subtitles": [{"line": 1, "in": 0.0, "out": 1.0, "text": "x"}],
                }
            ],
            "config": {"addLayers": {"stale": {"enabled": True}}},
        }

        try:
            with tempfile.TemporaryDirectory() as td:
                out_json = os.path.join(td, "out.json")
                with mock.patch.object(
                    mod, "convert_csv_to_json", return_value=mocked_data
                ):
                    rc = mod.main([in_path, out_json, "--layer-config", layer_cfg_path])
                self.assertEqual(rc, 0)
                with open(out_json, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                self.assertIn("config", payload)
                self.assertIn("addLayers", payload["config"])
                add_layers = payload["config"]["addLayers"]
                self.assertNotIn("stale", add_layers)
                self.assertIn("LAYER_NAME_CONFIG", add_layers)
                self.assertIn("TIMING_BEHAVIOR", add_layers)
                self.assertIn("TIMING_ITEM_SELECTOR", add_layers)
                self.assertIn("SKIP_COPY_CONFIG", add_layers)
        finally:
            try:
                os.remove(in_path)
            except Exception:
                pass
            try:
                os.remove(layer_cfg_path)
            except Exception:
                pass

    def test_missing_layer_config_warns_and_continues(self):
        csv_content = (
            "record_type;video_id;line;start;end;key;is_global;country_scope;metadata;GBL\n"
            "meta_global;;;;;briefVersion;Y;ALL;6;\n"
            "meta_global;;;;;fps;Y;ALL;25;\n"
            "meta_local;V;;;;title;N;ALL;T;\n"
            "sub;V;1;00:00:00:00;00:00:01:00;;;;;;;x\n"
        )
        in_path = tmp_file(".csv")
        with open(in_path, "w", encoding="utf-8") as f:
            f.write(csv_content)

        missing_layer_cfg = os.path.join(tempfile.gettempdir(), "missing_cfg_zzz.xlsx")
        if os.path.exists(missing_layer_cfg):
            os.remove(missing_layer_cfg)

        try:
            with tempfile.TemporaryDirectory() as td:
                out_json = os.path.join(td, "out.json")
                with mock.patch("sys.stdout", new_callable=io.StringIO) as out_buf:
                    with mock.patch("sys.stderr", new_callable=io.StringIO) as err_buf:
                        rc = mod.main(
                            [in_path, out_json, "--layer-config", missing_layer_cfg]
                        )
                self.assertEqual(rc, 0)
                self.assertTrue(os.path.isfile(out_json))
                self.assertIn(
                    f"Warning: failed to load layer config '{missing_layer_cfg}'",
                    err_buf.getvalue(),
                )
                self.assertIn(
                    "Conversion complete: Files written: 1, Errors: 1",
                    out_buf.getvalue(),
                )
                with open(out_json, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                self.assertTrue(isinstance(payload, dict))
                self.assertNotIn("config", payload)
        finally:
            try:
                os.remove(in_path)
            except Exception:
                pass

    def test_missing_media_config_warns_and_continues(self):
        csv_content = (
            "record_type;video_id;line;start;end;key;is_global;country_scope;metadata;GBL\n"
            "meta_global;;;;;briefVersion;Y;ALL;6;\n"
            "meta_global;;;;;fps;Y;ALL;25;\n"
            "meta_local;V;;;;title;N;ALL;T;\n"
            "sub;V;1;00:00:00:00;00:00:01:00;;;;;;;x\n"
        )
        in_path = tmp_file(".csv")
        with open(in_path, "w", encoding="utf-8") as f:
            f.write(csv_content)

        missing_media_cfg = os.path.join(tempfile.gettempdir(), "missing_media_zzz.csv")
        if os.path.exists(missing_media_cfg):
            os.remove(missing_media_cfg)

        try:
            with tempfile.TemporaryDirectory() as td:
                out_json = os.path.join(td, "out.json")
                with mock.patch("sys.stdout", new_callable=io.StringIO) as out_buf:
                    with mock.patch("sys.stderr", new_callable=io.StringIO) as err_buf:
                        rc = mod.main(
                            [in_path, out_json, "--media-config", missing_media_cfg]
                        )
                self.assertEqual(rc, 0)
                self.assertTrue(os.path.isfile(out_json))
                self.assertIn(
                    f"Warning: failed to load media config '{missing_media_cfg}'",
                    err_buf.getvalue(),
                )
                self.assertIn(
                    "Conversion complete: Files written: 1, Errors: 1",
                    out_buf.getvalue(),
                )
        finally:
            try:
                os.remove(in_path)
            except Exception:
                pass

    @unittest.skipUnless(Workbook is not None, "openpyxl is required for XLSX tests")
    def test_layer_config_converter_unavailable_uses_summary_path(self):
        csv_content = (
            "record_type;video_id;line;start;end;key;is_global;country_scope;metadata;GBL\n"
            "meta_global;;;;;briefVersion;Y;ALL;6;\n"
            "meta_global;;;;;fps;Y;ALL;25;\n"
            "meta_local;V;;;;title;N;ALL;T;\n"
            "sub;V;1;00:00:00:00;00:00:01:00;;;;;;;x\n"
        )
        in_path = tmp_file(".csv")
        with open(in_path, "w", encoding="utf-8") as f:
            f.write(csv_content)

        layer_cfg_path = tmp_file(".xlsx")
        self._write_layer_config_xlsx(layer_cfg_path)

        try:
            with tempfile.TemporaryDirectory() as td:
                out_json = os.path.join(td, "out.json")
                with mock.patch.object(mod, "layercfg_convert_workbook", None):
                    with mock.patch("sys.stdout", new_callable=io.StringIO) as out_buf:
                        with mock.patch(
                            "sys.stderr", new_callable=io.StringIO
                        ) as err_buf:
                            rc = mod.main(
                                [in_path, out_json, "--layer-config", layer_cfg_path]
                            )
                self.assertEqual(rc, 1)
                self.assertFalse(os.path.exists(out_json))
                self.assertIn(
                    "Layer config converter not available; cannot process --layer-config",
                    err_buf.getvalue(),
                )
                self.assertIn(
                    "Conversion complete: Files written: 0, Errors: 1",
                    out_buf.getvalue(),
                )
        finally:
            try:
                os.remove(in_path)
            except Exception:
                pass
            try:
                os.remove(layer_cfg_path)
            except Exception:
                pass

    @unittest.skipUnless(Workbook is not None, "openpyxl is required for XLSX tests")
    def test_layer_config_conversion_failure_uses_summary_path(self):
        csv_content = (
            "record_type;video_id;line;start;end;key;is_global;country_scope;metadata;GBL\n"
            "meta_global;;;;;briefVersion;Y;ALL;6;\n"
            "meta_global;;;;;fps;Y;ALL;25;\n"
            "meta_local;V;;;;title;N;ALL;T;\n"
            "sub;V;1;00:00:00:00;00:00:01:00;;;;;;;x\n"
        )
        in_path = tmp_file(".csv")
        with open(in_path, "w", encoding="utf-8") as f:
            f.write(csv_content)

        layer_cfg_path = tmp_file(".xlsx")
        self._write_layer_config_xlsx(layer_cfg_path)

        try:
            with tempfile.TemporaryDirectory() as td:
                out_json = os.path.join(td, "out.json")
                with mock.patch.object(
                    mod,
                    "layercfg_convert_workbook",
                    side_effect=ValueError("boom"),
                ):
                    with mock.patch("sys.stdout", new_callable=io.StringIO) as out_buf:
                        with mock.patch(
                            "sys.stderr", new_callable=io.StringIO
                        ) as err_buf:
                            rc = mod.main(
                                [in_path, out_json, "--layer-config", layer_cfg_path]
                            )
                self.assertEqual(rc, 1)
                self.assertFalse(os.path.exists(out_json))
                self.assertIn(
                    f"Failed to load layer config '{layer_cfg_path}': boom",
                    err_buf.getvalue(),
                )
                self.assertIn(
                    "Conversion complete: Files written: 0, Errors: 1",
                    out_buf.getvalue(),
                )
        finally:
            try:
                os.remove(in_path)
            except Exception:
                pass
            try:
                os.remove(layer_cfg_path)
            except Exception:
                pass


if __name__ == "__main__":
    unittest.main(verbosity=2)
